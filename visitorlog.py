import tkinter as tk
from tkinter import ttk, messagebox, filedialog, font as tkfont
import sqlite3
from datetime import datetime
import re
import os
import sys
import shutil

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# ============================================================
#  COLOR PALETTE  –  Modern red/white theme matching screenshot
# ============================================================
C_RED_DARK    = "#8B0000"
C_RED_MID     = "#B22222"
C_RED_ACCENT  = "#CC0000"
C_RED_LIGHT   = "#F5CCCC"
C_RED_BG      = "#FFE8E8"
C_WHITE       = "#FFFFFF"
C_OFF_WHITE   = "#F8F9FA"
C_GOLD        = "#D4A017"
C_TEXT_DARK   = "#2C0000"
C_TEXT_MID    = "#7A3030"
C_TEXT_LIGHT  = "#F2CECE"
C_SIDEBAR_BG  = "#8B0000"
C_SIDEBAR_HOVER = "#6B0000"
C_NAV_ACTIVE  = "#FFFFFF"
C_NAV_ACTIVE_BG = "#A00000"
C_BORDER      = "#E0C0C0"
C_INPUT_BG    = "#FFFFFF"
C_LABEL       = "#5A1A1A"
C_GREEN       = "#006633"

# ============================================================
#  BACKUP SYSTEM  –  Offline only (AppData\Local)
# ============================================================

# Base folder in AppData\Local\BarangayVisitorLog — always writable on any PC
APP_DATA_DIR = os.path.join(
    os.environ.get("LOCALAPPDATA", os.path.expanduser("~")),
    "BarangayVisitorLog"
)
BACKUP_DIR    = os.path.join(APP_DATA_DIR, "backups")
DB_PATH       = os.path.join(APP_DATA_DIR, "barangay_visitors.db")
MAX_BACKUPS   = 7
BACKUP_INTERVAL_MS = 3 * 60 * 60 * 1000   # check every 3 hours
_last_backup_time = None   # track when we last backed up

# Make sure the folder exists right away
os.makedirs(APP_DATA_DIR, exist_ok=True)


def get_backup_dir():
    """Return the backup folder path, creating it if needed."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    return BACKUP_DIR


def auto_backup_database(silent=True):
    """
    Copy barangay_visitors.db into AppData\\Local\\BarangayVisitorLog\\backups\\
    with a timestamp filename.  Keeps only the MAX_BACKUPS most-recent files.
    Returns (success: bool, message: str).
    """
    if not os.path.exists(DB_PATH):
        return False, "Database file not found – nothing to back up."

    try:
        backup_dir = get_backup_dir()
        timestamp  = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_name = f"barangay_visitors_backup_{timestamp}.db"
        backup_path = os.path.join(backup_dir, backup_name)

        shutil.copy2(DB_PATH, backup_path)

        # Prune old backups – keep only the newest MAX_BACKUPS
        all_backups = sorted(
            [f for f in os.listdir(backup_dir) if f.endswith(".db")],
            reverse=True
        )
        for old in all_backups[MAX_BACKUPS:]:
            try:
                os.remove(os.path.join(backup_dir, old))
            except Exception:
                pass

        msg = f"Backup saved:\n{backup_path}"
        if not silent:
            messagebox.showinfo("Backup Successful", msg)
        return True, msg

    except Exception as e:
        msg = f"Backup failed: {e}"
        if not silent:
            messagebox.showerror("Backup Failed", msg)
        return False, msg


def _get_db_row_count():
    """Return current number of visitor rows, or -1 if DB doesn't exist."""
    try:
        conn = sqlite3.connect(DB_PATH)
        count = conn.execute('SELECT COUNT(*) FROM visitors').fetchone()[0]
        conn.close()
        return count
    except Exception:
        return -1


def schedule_auto_backup(root):
    """
    Every 3 hours, only backs up if the visitor count changed since last backup.
    Keeps the backup folder from piling up with identical files.
    """
    global _last_backup_time
    current_count = _get_db_row_count()
    if current_count != _last_backup_time:  # _last_backup_time stores last backed-up row count
        ok, _ = auto_backup_database(silent=True)
        if ok:
            _last_backup_time = current_count
    root.after(BACKUP_INTERVAL_MS, lambda: schedule_auto_backup(root))


def get_backup_list():
    """Return a list of (filename, full_path, size_kb, modified_str) for all backups."""
    backup_dir = get_backup_dir()
    results = []
    try:
        files = sorted(
            [f for f in os.listdir(backup_dir) if f.endswith(".db")],
            reverse=True
        )
        for f in files:
            full = os.path.join(backup_dir, f)
            size_kb = os.path.getsize(full) / 1024
            modified = datetime.fromtimestamp(os.path.getmtime(full)).strftime("%Y-%m-%d  %I:%M %p")
            results.append((f, full, size_kb, modified))
    except Exception:
        pass
    return results


def restore_backup(backup_path, parent=None):
    """Overwrite the live DB with a chosen backup after confirmation."""
    if not os.path.exists(backup_path):
        messagebox.showerror("Error", "Backup file not found.", parent=parent)
        return False
    confirmed = messagebox.askyesno(
        "Restore Backup",
        "This will REPLACE the current database with the selected backup.\n\n"
        "A safety copy of the current database will be saved first.\n\n"
        "Continue?",
        parent=parent
    )
    if not confirmed:
        return False
    # Save current DB as safety copy first
    auto_backup_database(silent=True)
    try:
        shutil.copy2(backup_path, DB_PATH)
        messagebox.showinfo("Restore Successful",
                            "Database restored successfully!\n"
                            "Please restart the application.",
                            parent=parent)
        return True
    except Exception as e:
        messagebox.showerror("Restore Failed", f"Could not restore backup:\n{e}", parent=parent)
        return False


# ============================================================
#  DATABASE FUNCTIONS
# ============================================================

def create_database():
    auto_backup_database(silent=True)   # backup on every app launch
    connection = sqlite3.connect(DB_PATH)
    cursor = connection.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS visitors (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name       TEXT NOT NULL,
            contact         TEXT NOT NULL,
            address         TEXT NOT NULL,
            purpose         TEXT NOT NULL,
            visit_date      TEXT NOT NULL,
            visit_time      TEXT NOT NULL,
            person_to_see   TEXT DEFAULT ''
        )
    """)
    connection.commit()
    cursor.execute("PRAGMA table_info(visitors)")
    cols = [row[1] for row in cursor.fetchall()]
    for col, default in [('address', ''), ('person_to_see', '')]:
        if col not in cols:
            cursor.execute(f"ALTER TABLE visitors ADD COLUMN {col} TEXT DEFAULT '{default}'")
            connection.commit()
    connection.close()

def insert_visitor(full_name, contact, address, purpose, visit_date, visit_time, person_to_see=""):
    connection = sqlite3.connect(DB_PATH)
    cursor = connection.cursor()
    person_to_see = person_to_see.strip().title() if person_to_see else ""
    cursor.execute("""
        INSERT INTO visitors (full_name, contact, address, purpose, visit_date, visit_time, person_to_see)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (full_name, contact, address, purpose, visit_date, visit_time, person_to_see))
    connection.commit()
    connection.close()

def get_all_visitors(search=""):
    connection = sqlite3.connect(DB_PATH)
    cursor = connection.cursor()
    base = "SELECT id, full_name, contact, address, purpose, person_to_see, visit_date, visit_time FROM visitors"
    if search:
        like = "%" + search + "%"
        cursor.execute(base + " WHERE full_name LIKE ? OR contact LIKE ? OR purpose LIKE ? OR address LIKE ? ORDER BY id DESC",
                       (like, like, like, like))
    else:
        cursor.execute(base + " ORDER BY id DESC")
    results = cursor.fetchall()
    connection.close()
    return results

def update_visitor(visitor_id, full_name, contact, address, purpose, visit_date, visit_time, person_to_see=""):
    connection = sqlite3.connect(DB_PATH)
    cursor = connection.cursor()
    person_to_see = person_to_see.strip().title() if person_to_see else ""
    cursor.execute("""
        UPDATE visitors SET full_name=?, contact=?, address=?, purpose=?, visit_date=?, visit_time=?, person_to_see=? WHERE id=?
    """, (full_name, contact, address, purpose, visit_date, visit_time, person_to_see, visitor_id))
    connection.commit()
    connection.close()

def count_today():
    connection = sqlite3.connect(DB_PATH)
    cursor = connection.cursor()
    today = datetime.now().strftime("%Y-%m-%d")
    cursor.execute("SELECT COUNT(*) FROM visitors WHERE visit_date=?", (today,))
    total = cursor.fetchone()[0]
    connection.close()
    return total

def count_all():
    connection = sqlite3.connect(DB_PATH)
    cursor = connection.cursor()
    cursor.execute("SELECT COUNT(*) FROM visitors")
    total = cursor.fetchone()[0]
    connection.close()
    return total

# ============================================================
#  VALIDATION
# ============================================================

def is_valid_contact(contact):
    return bool(re.match(r"^(09|\+639)\d{9}$", contact))

def is_valid_date(date_text):
    try:
        datetime.strptime(date_text, "%Y-%m-%d"); return True
    except ValueError:
        return False

def is_valid_time(time_text):
    try:
        if not re.match(r"^\d{1,2}:\d{2}\s(AM|PM)$", time_text, re.IGNORECASE): return False
        datetime.strptime(time_text, "%I:%M %p"); return True
    except ValueError:
        return False

# ============================================================
#  ADMIN PASSWORD
# ============================================================

ADMIN_PASSWORD_FILE = os.path.join(APP_DATA_DIR, "admin_password.txt")

def get_admin_password():
    try:
        with open(ADMIN_PASSWORD_FILE, "r", encoding="utf-8") as f:
            return f.read().strip() or "admin123"
    except Exception:
        return "admin123"


def resource_path(relative_path):
    """Return absolute path to resource, works for dev and for PyInstaller bundle."""
    try:
        base = sys._MEIPASS
    except Exception:
        base = os.path.dirname(__file__)
    return os.path.join(base, relative_path)

def save_admin_password(new_password):
    try:
        with open(ADMIN_PASSWORD_FILE, "w", encoding="utf-8") as f:
            f.write(new_password)
        return True
    except Exception:
        return False

# ============================================================
#  STYLED INPUT HELPER
# ============================================================

def make_input_field(parent, label_text, var, placeholder="", icon="", width=None, is_combo=False, combo_values=None):
    """Creates a labeled input with icon, border styling."""
    frame = tk.Frame(parent, bg=C_WHITE)
    
    # Label
    lbl_frame = tk.Frame(frame, bg=C_WHITE)
    lbl_frame.pack(fill="x", pady=(0, 4))
    if icon:
        tk.Label(lbl_frame, text=icon, font=("Segoe UI", 9),
                 bg=C_WHITE, fg=C_RED_ACCENT).pack(side="left", padx=(0, 4))
    tk.Label(lbl_frame, text=label_text,
             font=("Segoe UI", 8, "bold"),
             bg=C_WHITE, fg=C_LABEL).pack(side="left")
    
    # Input container with border
    input_container = tk.Frame(frame, bg=C_INPUT_BG,
                                highlightbackground=C_BORDER,
                                highlightcolor=C_RED_ACCENT,
                                highlightthickness=1)
    input_container.pack(fill="x")
    
    if is_combo and combo_values:
        style = ttk.Style()
        style.configure("Card.TCombobox",
                        fieldbackground=C_INPUT_BG,
                        background=C_WHITE,
                        foreground=C_TEXT_DARK,
                        selectbackground=C_RED_LIGHT,
                        selectforeground=C_TEXT_DARK,
                        borderwidth=0)
        style.map("Card.TCombobox",
                  fieldbackground=[("readonly", C_INPUT_BG)],
                  background=[("active", C_RED_LIGHT)])
        widget = ttk.Combobox(input_container, textvariable=var,
                              values=combo_values, state="readonly",
                              font=("Segoe UI", 10), style="Card.TCombobox")
        widget.pack(fill="x", ipady=7, padx=2)
    else:
        widget = tk.Entry(input_container, textvariable=var,
                          font=("Segoe UI", 10),
                          bg=C_INPUT_BG, fg=C_TEXT_DARK,
                          insertbackground=C_RED_ACCENT,
                          relief="flat", bd=0)
        widget.pack(fill="x", ipady=8, padx=10)
    
    def on_focus_in(e):
        input_container.config(highlightbackground=C_RED_ACCENT, highlightthickness=2)
    def on_focus_out(e):
        input_container.config(highlightbackground=C_BORDER, highlightthickness=1)
    widget.bind("<FocusIn>", on_focus_in)
    widget.bind("<FocusOut>", on_focus_out)
    
    return frame, widget


def create_rounded_card_container(parent, width, height, bg, fill, outline, radius=24, padding=12):
    """Create a rounded card background and inner content frame."""
    try:
        from PIL import Image, ImageTk, ImageDraw
        # Start fully transparent so corners outside the rounded rect are see-through
        img = Image.new("RGBA", (width, height), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        draw.rounded_rectangle((0, 0, width, height), radius=radius, fill=fill, outline=outline, width=1)
        photo = ImageTk.PhotoImage(img)
        # Canvas bg must match the parent's bg so transparency blends correctly
        parent_bg = parent["bg"]
        canvas = tk.Canvas(parent, width=width, height=height,
                           bg=parent_bg, highlightthickness=0)
        canvas.image_ref = photo
        canvas.create_image(0, 0, anchor="nw", image=photo)
        inner = tk.Frame(canvas, bg=fill)
        canvas.create_window(padding, padding, anchor="nw", window=inner,
                             width=width - 2 * padding, height=height - 2 * padding)
        return canvas, inner
    except Exception:
        frame = tk.Frame(parent, bg=fill,
                         highlightbackground=outline,
                         highlightthickness=1)
        return frame, frame


def configure_button_styles():
    style = ttk.Style()
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure("RoundedAccent.TButton",
                    font=("Segoe UI", 10, "bold"),
                    foreground=C_WHITE,
                    background=C_RED_ACCENT,
                    borderwidth=0,
                    focusthickness=0,
                    focuscolor="",
                    padding=(14, 8))
    style.map("RoundedAccent.TButton",
              background=[("active", C_RED_DARK), ("disabled", C_RED_MID)])
    style.configure("RoundedMid.TButton",
                    font=("Segoe UI", 10, "bold"),
                    foreground=C_WHITE,
                    background=C_RED_MID,
                    borderwidth=0,
                    focusthickness=0,
                    focuscolor="",
                    padding=(12, 8))
    style.map("RoundedMid.TButton",
              background=[("active", C_RED_ACCENT), ("disabled", C_RED_MID)])
    style.configure("RoundedGreen.TButton",
                    font=("Segoe UI", 10, "bold"),
                    foreground=C_WHITE,
                    background=C_GREEN,
                    borderwidth=0,
                    focusthickness=0,
                    focuscolor="",
                    padding=(12, 8))
    style.map("RoundedGreen.TButton",
              background=[("active", "#005522"), ("disabled", "#006633")])


def load_icon_image_file(filename, size=18):
    path = resource_path(os.path.join("icon", filename))
    if not os.path.exists(path):
        return None
    try:
        from PIL import Image, ImageTk
        img = Image.open(path)
        if size:
            img = img.resize((size, size), Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception:
        try:
            return tk.PhotoImage(file=path)
        except Exception:
            return None


class RoundedButton(tk.Canvas):
    def __init__(self, parent, text="", icon=None, compound="left",
                 fg=C_WHITE, bg=C_RED_ACCENT, activebackground=C_RED_DARK,
                 font=("Segoe UI", 10, "bold"), command=None,
                 radius=16, padding=(18, 10), min_width=0, cursor="hand2", **kwargs):
        parent_bg = parent["bg"] if "bg" in parent.keys() else C_OFF_WHITE
        self.text = text
        self.icon = icon
        self.compound = compound
        self.fg = fg
        self.bg_color = bg
        self.active_bg = activebackground
        self.font = font
        self.command = command
        self.radius = radius
        self.padding_x, self.padding_y = padding
        self.min_width = min_width
        self._hover = False
        self._bg_image = None
        self._icon_ref = icon
        self._font = tkfont.Font(font=self.font)

        width = max(self._estimate_width(), self.min_width)
        height = max(self._estimate_height(), 36)
        super().__init__(parent, width=width, height=height,
                         bg=parent_bg, highlightthickness=0,
                         bd=0, cursor=cursor, **kwargs)

        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
        self.bind("<Button-1>", self._on_click)
        self.bind("<Configure>", self._on_configure)

        self._draw_button()

    def _estimate_width(self):
        text_width = self._font.measure(self.text)
        icon_width = self.icon.width() if self.icon else 0
        spacing = 8 if self.icon else 0
        return self.padding_x * 2 + text_width + icon_width + spacing

    def _estimate_height(self):
        text_height = self._font.metrics("linespace")
        icon_height = self.icon.height() if self.icon else 0
        return max(text_height, icon_height) + self.padding_y * 2

    def _create_rounded_image(self, width, height, color):
        try:
            from PIL import Image, ImageDraw, ImageTk
            img = Image.new("RGBA", (width, height), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)
            draw.rounded_rectangle((0, 0, width, height), radius=self.radius,
                                   fill=color)
            return ImageTk.PhotoImage(img)
        except Exception:
            return None

    def _draw_button(self):
        self.delete("all")
        width = max(self.winfo_width(), self._estimate_width(), self.min_width)
        height = max(self.winfo_height(), self._estimate_height())
        tk.Canvas.config(self, width=width, height=height)

        bg_color = self.active_bg if self._hover else self.bg_color
        self._bg_image = self._create_rounded_image(width, height, bg_color)
        if self._bg_image:
            self.create_image(0, 0, image=self._bg_image, anchor="nw")

        center_y = height // 2
        if self.icon and self.compound == "left":
            text_width = self._font.measure(self.text)
            icon_width = self.icon.width()
            total_width = icon_width + 8 + text_width
            content_x = max(self.padding_x, (width - total_width) // 2)
            self.create_image(content_x, center_y, image=self.icon, anchor="w")
            content_x += icon_width + 8
            self.create_text(content_x, center_y, text=self.text,
                             anchor="w", fill=self.fg, font=self.font)
        else:
            self.create_text(width // 2, center_y, text=self.text,
                             anchor="center", fill=self.fg, font=self.font)

    def _on_enter(self, event=None):
        self._hover = True
        self._draw_button()

    def _on_leave(self, event=None):
        self._hover = False
        self._draw_button()

    def _on_click(self, event=None):
        if callable(self.command):
            self.command()

    def config(self, **kwargs):
        if "command" in kwargs:
            self.command = kwargs.pop("command")
        if "text" in kwargs:
            self.text = kwargs.pop("text")
        if "icon" in kwargs:
            self.icon = kwargs.pop("icon")
            self._icon_ref = self.icon
        if kwargs:
            super().config(**kwargs)
        self._draw_button()
        return self

    configure = config

    def _on_configure(self, event):
        if event.width != self.winfo_width() or event.height != self.winfo_height():
            self._draw_button()


# ============================================================
#  LOGIN WINDOW
# ============================================================

def open_login_window(parent=None):
    login = tk.Toplevel(parent)
    login.title("Admin Login")
    width, height = 380, 280
    login.geometry(f"{width}x{height}")
    login.resizable(False, False)
    login.configure(bg=C_WHITE)
    login.update_idletasks()
    try:
        if parent:
            x = parent.winfo_rootx() + (parent.winfo_width() - width) // 2
            y = parent.winfo_rooty() + (parent.winfo_height() - height) // 2
        else:
            x = (login.winfo_screenwidth() - width) // 2
            y = (login.winfo_screenheight() - height) // 2
    except Exception:
        x, y = 100, 100
    login.geometry(f"{width}x{height}+{x}+{y}")
    login.grab_set()

    # Red top bar
    top = tk.Frame(login, bg=C_RED_ACCENT, height=8)
    top.pack(fill="x")

    # Header
    header = tk.Frame(login, bg=C_RED_DARK)
    header.pack(fill="x")
    tk.Label(header, text="🔒  Administrator Login",
             font=("Georgia", 13, "bold"),
             bg=C_RED_DARK, fg=C_WHITE, pady=16).pack()

    body = tk.Frame(login, bg=C_WHITE, padx=40, pady=20)
    body.pack(fill="both", expand=True)

    tk.Label(body, text="Password",
             font=("Segoe UI", 9, "bold"),
             bg=C_WHITE, fg=C_LABEL).pack(anchor="w", pady=(0, 4))

    pw_container = tk.Frame(body, bg=C_INPUT_BG,
                             highlightbackground=C_BORDER,
                             highlightcolor=C_RED_ACCENT,
                             highlightthickness=1)
    pw_container.pack(fill="x")
    password_var = tk.StringVar()
    entry = tk.Entry(pw_container, textvariable=password_var, show="●",
                     font=("Segoe UI", 11), bg=C_INPUT_BG, fg=C_TEXT_DARK,
                     insertbackground=C_RED_ACCENT, relief="flat", bd=0)
    entry.pack(fill="x", ipady=9, padx=10)
    entry.focus()

    def on_focus_in(e): pw_container.config(highlightbackground=C_RED_ACCENT, highlightthickness=2)
    def on_focus_out(e): pw_container.config(highlightbackground=C_BORDER, highlightthickness=1)
    entry.bind("<FocusIn>", on_focus_in)
    entry.bind("<FocusOut>", on_focus_out)

    result = {"success": False}

    def check_password():
        if password_var.get() == get_admin_password():
            result["success"] = True
            login.destroy()
        else:
            messagebox.showerror("Wrong Password", "Incorrect password. Try again.", parent=login)
            password_var.set("")

    btn = RoundedButton(body, text="LOGIN",
                         bg=C_RED_ACCENT,
                         activebackground=C_RED_DARK,
                         min_width=260,
                         command=check_password)
    btn.pack(pady=(16, 0), anchor="center")

    entry.bind("<Return>", lambda e: check_password())
    login.wait_window()
    return result["success"]


def open_change_password_window(parent=None):
    dialog = tk.Toplevel(parent)
    dialog.title("Change Admin Password")
    width, height = 420, 360
    dialog.geometry(f"{width}x{height}")
    dialog.resizable(False, False)
    dialog.configure(bg=C_WHITE)
    dialog.update_idletasks()
    try:
        if parent:
            x = parent.winfo_rootx() + (parent.winfo_width() - width) // 2
            y = parent.winfo_rooty() + (parent.winfo_height() - height) // 2
        else:
            x = (dialog.winfo_screenwidth() - width) // 2
            y = (dialog.winfo_screenheight() - height) // 2
    except Exception:
        x, y = 100, 100
    dialog.geometry(f"{width}x{height}+{x}+{y}")
    dialog.grab_set()

    top = tk.Frame(dialog, bg=C_RED_ACCENT, height=8)
    top.pack(fill="x")

    header = tk.Frame(dialog, bg=C_RED_DARK)
    header.pack(fill="x", pady=(0, 8))
    lock_icon = load_icon_image_file("lock.png", 24)
    if lock_icon:
        icon_lbl = tk.Label(header, image=lock_icon, bg=C_RED_DARK)
        icon_lbl._img = lock_icon
        icon_lbl.pack(side="left", padx=(20, 8), pady=12)
    tk.Label(header, text="Change Admin Password",
             font=("Georgia", 13, "bold"),
             bg=C_RED_DARK, fg=C_WHITE, pady=16).pack(side="left", padx=(0, 16))

    body = tk.Frame(dialog, bg=C_WHITE, padx=30, pady=20)
    body.pack(fill="both", expand=True)

    current_var = tk.StringVar()
    new_var = tk.StringVar()

    for label_text, var in [("Current Password", current_var), ("New Password", new_var)]:
        lbl = tk.Label(body, text=label_text,
                       font=("Segoe UI", 9, "bold"),
                       bg=C_WHITE, fg=C_LABEL)
        lbl.pack(anchor="w", pady=(0, 6))
        cont = tk.Frame(body, bg=C_INPUT_BG,
                        highlightbackground=C_BORDER, highlightthickness=1)
        cont.pack(fill="x", pady=(0, 14))
        ent = tk.Entry(cont, textvariable=var, show="●",
                       font=("Segoe UI", 10), bg=C_INPUT_BG, fg=C_TEXT_DARK,
                       insertbackground=C_RED_ACCENT, relief="flat", bd=0)
        ent.pack(fill="x", ipady=8, padx=10)
        ent.bind("<FocusIn>", lambda e, c=cont: c.config(highlightbackground=C_RED_ACCENT, highlightthickness=2))
        ent.bind("<FocusOut>", lambda e, c=cont: c.config(highlightbackground=C_BORDER, highlightthickness=1))

    result = {"success": False}

    def save_new_password():
        if current_var.get() != get_admin_password():
            messagebox.showerror("Error", "Current password is incorrect.", parent=dialog)
            return
        if len(new_var.get()) < 4:
            messagebox.showerror("Error", "New password must be at least 4 characters.", parent=dialog)
            return
        if save_admin_password(new_var.get()):
            messagebox.showinfo("Success", "Password updated successfully.", parent=dialog)
            result["success"] = True
            dialog.destroy()
        else:
            messagebox.showerror("Error", "Failed to save new password.", parent=dialog)

    btn = RoundedButton(body, text="Update Password",
                         bg=C_RED_ACCENT,
                         activebackground=C_RED_DARK,
                         command=save_new_password)
    btn.pack(pady=(10, 0), anchor="center")

    dialog.bind("<Return>", lambda e: save_new_password())
    dialog.wait_window()
    return result["success"]

# ============================================================
#  SIDEBAR NAV ITEM
# ============================================================

def make_nav_item(sidebar, icon, text, active=False, command=None):
    bg = C_NAV_ACTIVE_BG if active else C_SIDEBAR_BG
    frame = tk.Frame(sidebar, bg=bg, cursor="hand2")
    frame.pack(fill="x", padx=0, pady=1)

    # Left accent bar for active
    if active:
        tk.Frame(frame, bg=C_GOLD, width=4).pack(side="left", fill="y")

    inner = tk.Frame(frame, bg=bg)
    inner.pack(side="left", fill="x", expand=True, padx=(12 if not active else 8), pady=10)

    tk.Label(inner, text=icon, font=("Segoe UI", 11),
             bg=bg, fg=C_WHITE).pack(side="left")
    tk.Label(inner, text=f"  {text}", font=("Segoe UI", 10, "bold" if active else "normal"),
             bg=bg, fg=C_WHITE).pack(side="left")

    def on_enter(e):
        if not active:
            frame.config(bg=C_SIDEBAR_HOVER)
            inner.config(bg=C_SIDEBAR_HOVER)
            for w in inner.winfo_children():
                w.config(bg=C_SIDEBAR_HOVER)
    def on_leave(e):
        if not active:
            frame.config(bg=C_SIDEBAR_BG)
            inner.config(bg=C_SIDEBAR_BG)
            for w in inner.winfo_children():
                w.config(bg=C_SIDEBAR_BG)

    frame.bind("<Enter>", on_enter)
    frame.bind("<Leave>", on_leave)
    inner.bind("<Enter>", on_enter)
    inner.bind("<Leave>", on_leave)
    for w in inner.winfo_children():
        w.bind("<Enter>", on_enter)
        w.bind("<Leave>", on_leave)

    if command:
        for widget in [frame, inner] + list(inner.winfo_children()):
            widget.bind("<Button-1>", lambda e: command())

    return frame

# ============================================================
#  USER SCREEN  –  Modern Visitor Entry Form
# ============================================================

def show_user_screen(root, main_frame, sidebar_nav=None):
    for w in main_frame.winfo_children():
        w.destroy()

    # Update sidebar nav if available
    if sidebar_nav:
        sidebar_nav["update"]("visitor_entry")

    # ── TOP BAR ──────────────────────────────────────────────
    topbar = tk.Frame(main_frame, bg=C_RED_DARK, height=56)
    topbar.pack(fill="x")
    topbar.pack_propagate(False)

    # Logo area in topbar
    left_top = tk.Frame(topbar, bg=C_RED_DARK)
    left_top.pack(side="left", padx=14, fill="y")

    logo_dir = resource_path("logo")
    logo_path = resource_path(os.path.join("logo", "logo.png"))
    _top_logo = [None]

    def load_png_image(path, size=None):
        if not os.path.exists(path):
            return None
        try:
            from PIL import Image, ImageTk
            img = Image.open(path)
            if size:
                img = img.resize((size, size), Image.LANCZOS)
            return ImageTk.PhotoImage(img)
        except Exception:
            try:
                return tk.PhotoImage(file=path)
            except Exception:
                return None

    top_logo_img = load_png_image(logo_path, 50)
    if top_logo_img:
        _top_logo[0] = top_logo_img
        lbl = tk.Label(left_top, image=_top_logo[0], bg=C_RED_DARK)
        lbl.pack(side="left", padx=(0, 8))
        lbl._img = _top_logo[0]

    icon_dir = resource_path("icon")
    _form_icons = {}
    def load_icon_image(filename, size=18):
        icon_img = load_icon_image_file(filename, size)
        if icon_img:
            _form_icons[filename] = icon_img
        return icon_img

    title_frame = tk.Frame(left_top, bg=C_RED_DARK)
    title_frame.pack(side="left", fill="y")
    tk.Label(title_frame, text="BARANGAY SAN ANDRES",
             font=("Georgia", 11, "bold"),
             bg=C_RED_DARK, fg=C_WHITE).pack(anchor="w")
    tk.Label(title_frame, text="Visitor Log System",
             font=("Segoe UI", 8),
             bg=C_RED_DARK, fg=C_TEXT_LIGHT).pack(anchor="w")

    # Marquee strip in center
    marquee_frame = tk.Frame(topbar, bg=C_RED_DARK)
    marquee_frame.pack(side="left", expand=True, fill="y")

    marquee_inner = tk.Frame(marquee_frame, bg=C_RED_DARK)
    marquee_inner.place(relx=0.5, rely=0.5, anchor="center")

    # Decorative dots
    tk.Label(marquee_inner, text="— ✦ —  Mabuhay! Please register your visit.  — ✦ —",
             font=("Segoe UI", 9, "italic"),
             bg=C_RED_DARK, fg=C_TEXT_LIGHT).pack()

    # Admin login button
    admin_icon = load_icon_image("user.png", 18)
    admin_btn = RoundedButton(topbar, text="Admin Login",
                               icon=admin_icon, compound="left",
                               bg=C_RED_MID,
                               activebackground=C_RED_ACCENT,
                               min_width=220,
                               command=lambda: try_admin_login(root, main_frame, sidebar_nav))
    if admin_icon:
        admin_btn._icon_ref = admin_icon
    admin_btn.pack(side="right", padx=14, pady=8)

    # Gold accent line
    tk.Frame(main_frame, bg=C_GOLD, height=3).pack(fill="x")

    # ── SCROLLABLE CONTENT AREA ───────────────────────────────
    content_canvas = tk.Canvas(main_frame, bg=C_OFF_WHITE, highlightthickness=0)
    vsb = ttk.Scrollbar(main_frame, orient="vertical", command=content_canvas.yview)
    content_canvas.configure(yscrollcommand=vsb.set)
    vsb.pack(side="right", fill="y")
    content_canvas.pack(side="left", fill="both", expand=True)

    # Background image
    _bg_ref = [None]
    _bg_path = resource_path(os.path.join("logo", "barangay.png"))

    def _draw_bg():
        w = content_canvas.winfo_width()
        h = content_canvas.winfo_height()
        if w < 10 or h < 10:
            return
        content_canvas.delete("bg_layer")
        try:
            from PIL import Image, ImageTk, ImageEnhance
            img = Image.open(_bg_path).convert("RGBA")
            iw, ih = img.size
            scale = max(w / iw, h / ih)
            nw, nh = int(iw * scale), int(ih * scale)
            img = img.resize((nw, nh), Image.LANCZOS)
            left = (nw - w) // 2
            top = max(0, (nh - h) // 2)
            img = img.crop((left, top, left + w, top + h))
            rgb = ImageEnhance.Brightness(img.convert("RGB")).enhance(0.55)
            photo = ImageTk.PhotoImage(rgb)
            _bg_ref[0] = photo
            content_canvas.create_image(0, 0, anchor="nw", image=photo, tags="bg_layer")
            content_canvas.tag_lower("bg_layer")
        except Exception:
            content_canvas.configure(bg=C_OFF_WHITE)


    def _on_canvas_configure(e):
        _draw_bg()

    content_canvas.bind("<Configure>", _on_canvas_configure)
    content_canvas.after(150, _draw_bg)

    # Enable mousewheel scrolling
    def _on_mousewheel(e):
        content_canvas.yview_scroll(int(-1*(e.delta/120)), "units")
    content_canvas.bind_all("<MouseWheel>", _on_mousewheel)

    # ── CARD — draw rounded rect directly on content_canvas for true transparent corners ──
    CARD_W = 700
    CARD_MIN_H = 570
    CARD_RADIUS = 24

    _card_rect_id = [None]
    def _draw_card_rect(x, y, height):
        if _card_rect_id[0]:
            content_canvas.delete(_card_rect_id[0])
        x2, y2 = x + CARD_W, y + height
        r = CARD_RADIUS
        points = [
            x+r, y,  x2-r, y,
            x2, y,   x2, y+r,
            x2, y2-r, x2, y2,
            x2-r, y2, x+r, y2,
            x, y2,   x, y2-r,
            x, y+r,  x, y,
        ]
        _card_rect_id[0] = content_canvas.create_polygon(
            points, smooth=True,
            fill=C_WHITE, outline=C_BORDER, width=1,
            tags="card_bg"
        )
        content_canvas.tag_lower("card_bg", "card_window")

    card = tk.Frame(content_canvas, bg=C_WHITE)
    card_id = content_canvas.create_window(0, 0, anchor="nw", window=card,
                                           width=CARD_W,
                                           tags="card_window")

    def _center_card(e=None):
        content_canvas.update_idletasks()
        card_h = max(CARD_MIN_H, card.winfo_reqheight())
        cw = content_canvas.winfo_width()
        ch = content_canvas.winfo_height()
        x = max(0, (cw - CARD_W) // 2)
        y = max(20, (ch - card_h) // 2)
        content_canvas.coords(card_id, x, y)
        content_canvas.itemconfig(card_id, width=CARD_W, height=card_h)
        _draw_card_rect(x, y, card_h)
        scroll_h = max(ch, card_h + y * 2)
        content_canvas.configure(scrollregion=(0, 0, cw, scroll_h))
        content_canvas.tag_lower("bg_layer")

    content_canvas.bind("<Configure>", lambda e: [_on_canvas_configure(e), _center_card(e)])
    content_canvas.after(200, _center_card)

    # Red top accent
    tk.Frame(card, bg=C_RED_ACCENT, height=6).pack(fill="x")

    # ── CARD HEADER ──────────────────────────────────────────
    header_frame = tk.Frame(card, bg=C_WHITE)
    header_frame.pack(fill="x", padx=30, pady=(20, 0))

    # Left logo
    _card_logos = [None, None]

    def _load_logo(path, size=90):
        if not os.path.exists(path):
            return None
        try:
            from PIL import Image, ImageTk
            img = Image.open(path)
            # Use thumbnail to maintain aspect ratio while fitting within size x size box
            img.thumbnail((size, size), Image.LANCZOS)
            return ImageTk.PhotoImage(img)
        except Exception:
            try:
                return tk.PhotoImage(file=path)
            except Exception:
                return None

    logo1 = _load_logo(os.path.join(logo_dir, "stgo.png"), 112)
    logo2 = _load_logo(os.path.join(logo_dir, "logo.png"), 112)
    _card_logos[0], _card_logos[1] = logo1, logo2

    left_logo_lbl = tk.Label(header_frame, bg=C_WHITE)
    if logo1:
        left_logo_lbl.config(image=logo1)
        left_logo_lbl._img = logo1
    else:
        left_logo_lbl.config(text="🏛", font=("Segoe UI", 28), fg=C_RED_DARK)
    left_logo_lbl.pack(side="left")

    center_header = tk.Frame(header_frame, bg=C_WHITE)
    center_header.pack(side="left", expand=True)

    tk.Label(center_header, text="WELCOME TO",
             font=("Georgia", 10),
             bg=C_WHITE, fg=C_TEXT_MID).pack()
    tk.Label(center_header, text="BARANGAY SAN ANDRES",
             font=("Georgia", 18, "bold"),
             bg=C_WHITE, fg=C_RED_DARK).pack()
    tk.Label(center_header, text="We're glad you're here. Please fill in your details below.",
             font=("Segoe UI", 9),
             bg=C_WHITE, fg=C_TEXT_MID).pack(pady=(4, 0))

    # Gold divider with diamond
    div_frame = tk.Frame(center_header, bg=C_WHITE, height=20)
    div_frame.pack(fill="x", pady=(8, 0))
    div_canvas = tk.Canvas(div_frame, bg=C_WHITE, height=20, highlightthickness=0)
    div_canvas.pack(fill="x")
    def draw_divider(e=None):
        w = div_canvas.winfo_width()
        if w < 10: return
        mid = w // 2
        div_canvas.delete("all")
        div_canvas.create_line(mid - 80, 10, mid - 12, 10, fill=C_GOLD, width=2)
        div_canvas.create_oval(mid - 6, 6, mid + 6, 14, fill=C_GOLD, outline=C_GOLD)
        div_canvas.create_line(mid + 12, 10, mid + 80, 10, fill=C_GOLD, width=2)
    div_canvas.bind("<Configure>", lambda e: draw_divider())

    right_logo_lbl = tk.Label(header_frame, bg=C_WHITE)
    if logo2:
        right_logo_lbl.config(image=logo2)
        right_logo_lbl._img = logo2
    else:
        right_logo_lbl.config(text="SK", font=("Segoe UI", 22, "bold"), fg=C_RED_DARK)
    right_logo_lbl.pack(side="right")

    # ── FORM ─────────────────────────────────────────────────
    form_outer = tk.Frame(card, bg=C_WHITE)
    form_outer.pack(fill="x", padx=30, pady=(20, 0))

    # Thin separator
    tk.Frame(form_outer, bg=C_BORDER, height=1).pack(fill="x", pady=(0, 20))

    purpose_options = [
        "Business / Transaction",
        "Certificate Request",
        "Complaint / Concern",
        "Clearance Application",
        "Meeting / Appointment",
        "Social Service Inquiry",
        "Other",
    ]

    name_var          = tk.StringVar()
    contact_var       = tk.StringVar()
    address_var       = tk.StringVar()
    purpose_var       = tk.StringVar(value="Select purpose of visit")
    other_purpose_var = tk.StringVar()
    
    date_var          = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
    time_var          = tk.StringVar(value=datetime.now().strftime("%I:%M %p"))
    person_var        = tk.StringVar()
    
    # Register form variables to sync with the sidebar clock
    if sidebar_nav:
        # Access the parent scope's form_vars dictionary through root
        try:
            root.form_vars["date_var"] = date_var
            root.form_vars["time_var"] = time_var
        except Exception:
            pass

    # Row 1: Full Name + Contact
    row1 = tk.Frame(form_outer, bg=C_WHITE)
    row1.pack(fill="x", pady=(0, 16))

    name_frame = tk.Frame(row1, bg=C_WHITE)
    name_frame.pack(side="left", fill="x", expand=True, padx=(0, 12))

    name_label_frame = tk.Frame(name_frame, bg=C_WHITE)
    name_label_frame.pack(fill="x", pady=(0, 5))
    name_icon = load_icon_image("user.png", 18)
    if name_icon:
        name_icon_lbl = tk.Label(name_label_frame, image=name_icon, bg=C_WHITE)
        name_icon_lbl._img = name_icon
        name_icon_lbl.pack(side="left", padx=(0, 4))
    tk.Label(name_label_frame, text="FULL NAME *",
             font=("Segoe UI", 8, "bold"),
             bg=C_WHITE, fg=C_LABEL).pack(side="left")
    name_container = tk.Frame(name_frame, bg=C_INPUT_BG,
                               highlightbackground=C_BORDER,
                               highlightcolor=C_RED_ACCENT,
                               highlightthickness=1)
    name_container.pack(fill="x")
    name_entry = tk.Entry(name_container, textvariable=name_var,
                          font=("Segoe UI", 10), bg=C_INPUT_BG, fg=C_TEXT_DARK,
                          insertbackground=C_RED_ACCENT, relief="flat", bd=0)
    name_entry.pack(fill="x", ipady=9, padx=12)

    contact_frame = tk.Frame(row1, bg=C_WHITE)
    contact_frame.pack(side="left", fill="x", expand=True)

    contact_label_frame = tk.Frame(contact_frame, bg=C_WHITE)
    contact_label_frame.pack(fill="x", pady=(0, 5))
    contact_icon = load_icon_image("telephone-handle-silhouette.png", 18)
    if contact_icon:
        contact_icon_lbl = tk.Label(contact_label_frame, image=contact_icon, bg=C_WHITE)
        contact_icon_lbl._img = contact_icon
        contact_icon_lbl.pack(side="left", padx=(0, 4))
    tk.Label(contact_label_frame, text="CONTACT NUMBER *",
             font=("Segoe UI", 8, "bold"),
             bg=C_WHITE, fg=C_LABEL).pack(side="left")
    contact_container = tk.Frame(contact_frame, bg=C_INPUT_BG,
                                  highlightbackground=C_BORDER,
                                  highlightcolor=C_RED_ACCENT,
                                  highlightthickness=1)
    contact_container.pack(fill="x")
    contact_entry = tk.Entry(contact_container, textvariable=contact_var,
                              font=("Segoe UI", 10), bg=C_INPUT_BG, fg=C_TEXT_DARK,
                              insertbackground=C_RED_ACCENT, relief="flat", bd=0)
    contact_entry.pack(fill="x", ipady=9, padx=12)

    # Row 2: Address (full width)
    row2 = tk.Frame(form_outer, bg=C_WHITE)
    row2.pack(fill="x", pady=(0, 16))

    address_label_frame = tk.Frame(row2, bg=C_WHITE)
    address_label_frame.pack(fill="x", pady=(0, 5))
    address_icon = load_icon_image("home.png", 18)
    if address_icon:
        address_icon_lbl = tk.Label(address_label_frame, image=address_icon, bg=C_WHITE)
        address_icon_lbl._img = address_icon
        address_icon_lbl.pack(side="left", padx=(0, 4))
    tk.Label(address_label_frame, text="ADDRESS *",
             font=("Segoe UI", 8, "bold"),
             bg=C_WHITE, fg=C_LABEL).pack(side="left")
    addr_container = tk.Frame(row2, bg=C_INPUT_BG,
                               highlightbackground=C_BORDER,
                               highlightcolor=C_RED_ACCENT,
                               highlightthickness=1)
    addr_container.pack(fill="x")
    addr_entry = tk.Entry(addr_container, textvariable=address_var,
                           font=("Segoe UI", 10), bg=C_INPUT_BG, fg=C_TEXT_DARK,
                           insertbackground=C_RED_ACCENT, relief="flat", bd=0)
    addr_entry.pack(fill="x", ipady=9, padx=12)

    # Row 3: Purpose + Date
    row3 = tk.Frame(form_outer, bg=C_WHITE)
    row3.pack(fill="x", pady=(0, 16))

    purpose_frame = tk.Frame(row3, bg=C_WHITE)
    purpose_frame.pack(side="left", fill="x", expand=True, padx=(0, 12))

    purpose_label_frame = tk.Frame(purpose_frame, bg=C_WHITE)
    purpose_label_frame.pack(fill="x", pady=(0, 5))
    purpose_icon = load_icon_image("target.png", 18)
    if purpose_icon:
        purpose_icon_lbl = tk.Label(purpose_label_frame, image=purpose_icon, bg=C_WHITE)
        purpose_icon_lbl._img = purpose_icon
        purpose_icon_lbl.pack(side="left", padx=(0, 4))
    tk.Label(purpose_label_frame, text="PURPOSE OF VISIT *",
             font=("Segoe UI", 8, "bold"),
             bg=C_WHITE, fg=C_LABEL).pack(side="left")

    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Card.TCombobox",
                    fieldbackground=C_INPUT_BG,
                    background=C_WHITE,
                    foreground=C_TEXT_DARK,
                    selectbackground=C_RED_LIGHT,
                    selectforeground=C_TEXT_DARK,
                    arrowcolor=C_RED_ACCENT,
                    borderwidth=0)
    style.map("Card.TCombobox",
              fieldbackground=[("readonly", C_INPUT_BG)],
              background=[("active", C_RED_LIGHT)],
              arrowcolor=[("active", C_RED_DARK)])

    purpose_container = tk.Frame(purpose_frame, bg=C_INPUT_BG,
                                  highlightbackground=C_BORDER,
                                  highlightcolor=C_RED_ACCENT,
                                  highlightthickness=1)
    purpose_container.pack(fill="x")
    purpose_combo = ttk.Combobox(purpose_container, textvariable=purpose_var,
                                  values=purpose_options, state="readonly",
                                  font=("Segoe UI", 10), style="Card.TCombobox")
    purpose_combo.pack(fill="x", ipady=6, padx=2)

    date_frame = tk.Frame(row3, bg=C_WHITE)
    date_frame.pack(side="left", fill="x", expand=True)

    date_label_frame = tk.Frame(date_frame, bg=C_WHITE)
    date_label_frame.pack(fill="x", pady=(0, 5))
    date_icon = load_icon_image("calendar.png", 18)
    if date_icon:
        date_icon_lbl = tk.Label(date_label_frame, image=date_icon, bg=C_WHITE)
        date_icon_lbl._img = date_icon
        date_icon_lbl.pack(side="left", padx=(0, 4))
    tk.Label(date_label_frame, text="VISIT DATE *",
             font=("Segoe UI", 8, "bold"),
             bg=C_WHITE, fg=C_LABEL).pack(side="left")
    date_container = tk.Frame(date_frame, bg=C_INPUT_BG,
                               highlightbackground=C_BORDER,
                               highlightcolor=C_RED_ACCENT,
                               highlightthickness=1)
    date_container.pack(fill="x")
    date_entry = tk.Entry(date_container, textvariable=date_var,
                           font=("Segoe UI", 10), bg=C_INPUT_BG, fg=C_TEXT_DARK,
                           insertbackground=C_RED_ACCENT, relief="flat", bd=0,
                           state="readonly")
    date_entry.pack(fill="x", ipady=9, padx=12)

    # Row 4: "Other" purpose field (hidden by default)
    other_row = tk.Frame(form_outer, bg=C_WHITE, height=0)
    other_row.pack_propagate(False)
    other_inner = tk.Frame(other_row, bg=C_WHITE)
    other_inner.pack(fill="both", expand=True)
    tk.Label(other_inner, text="✏  PLEASE SPECIFY PURPOSE *",
             font=("Segoe UI", 8, "bold"),
             bg=C_WHITE, fg=C_LABEL).pack(anchor="w", pady=(0, 5))
    other_container = tk.Frame(other_inner, bg=C_INPUT_BG,
                                highlightbackground=C_BORDER,
                                highlightcolor=C_RED_ACCENT,
                                highlightthickness=1)
    other_container.pack(fill="x")
    other_entry = tk.Entry(other_container, textvariable=other_purpose_var,
                            font=("Segoe UI", 10), bg=C_INPUT_BG, fg=C_TEXT_DARK,
                            insertbackground=C_RED_ACCENT, relief="flat", bd=0)
    other_entry.pack(fill="x", ipady=9, padx=12)
    other_row.pack_forget()
    other_anim_job = [None]

    def _cancel_other_animation():
        if other_anim_job[0] is not None:
            other_row.after_cancel(other_anim_job[0])
            other_anim_job[0] = None

    def _animate_other(show, current_height=None):
        _cancel_other_animation()
        target = 88
        step = 12
        if current_height is None:
            current_height = other_row.winfo_height()
        if show:
            if current_height >= target:
                other_row.configure(height=target)
                other_anim_job[0] = None
                _center_card()
                return
            next_height = min(target, current_height + step)
            other_row.configure(height=next_height)
            _center_card()
            other_anim_job[0] = other_row.after(10, lambda: _animate_other(True, next_height))
        else:
            if current_height <= 0:
                other_row.pack_forget()
                other_row.configure(height=0)
                other_anim_job[0] = None
                _center_card()
                return
            next_height = max(0, current_height - step)
            other_row.configure(height=next_height)
            _center_card()
            other_anim_job[0] = other_row.after(10, lambda: _animate_other(False, next_height))

    # Row 5: Time + Person to see
    row5 = tk.Frame(form_outer, bg=C_WHITE)
    row5.pack(fill="x", pady=(0, 16))

    time_frame = tk.Frame(row5, bg=C_WHITE)
    time_frame.pack(side="left", fill="x", expand=True, padx=(0, 12))

    time_label_frame = tk.Frame(time_frame, bg=C_WHITE)
    time_label_frame.pack(fill="x", pady=(0, 5))
    time_icon = load_icon_image("clock.png", 18)
    if time_icon:
        time_icon_lbl = tk.Label(time_label_frame, image=time_icon, bg=C_WHITE)
        time_icon_lbl._img = time_icon
        time_icon_lbl.pack(side="left", padx=(0, 4))
    tk.Label(time_label_frame, text="VISIT TIME *",
             font=("Segoe UI", 8, "bold"),
             bg=C_WHITE, fg=C_LABEL).pack(side="left")
    time_container = tk.Frame(time_frame, bg=C_INPUT_BG,
                               highlightbackground=C_BORDER,
                               highlightcolor=C_RED_ACCENT,
                               highlightthickness=1)
    time_container.pack(fill="x")
    time_entry = tk.Entry(time_container, textvariable=time_var,
                           font=("Segoe UI", 10), bg=C_INPUT_BG, fg=C_TEXT_DARK,
                           insertbackground=C_RED_ACCENT, relief="flat", bd=0,
                           state="readonly")
    time_entry.pack(fill="x", ipady=9, padx=12)

    person_frame = tk.Frame(row5, bg=C_WHITE)
    person_frame.pack(side="left", fill="x", expand=True)

    person_label_frame = tk.Frame(person_frame, bg=C_WHITE)
    person_label_frame.pack(fill="x", pady=(0, 5))
    person_icon = load_icon_image("user.png", 18)
    if person_icon:
        person_icon_lbl = tk.Label(person_label_frame, image=person_icon, bg=C_WHITE)
        person_icon_lbl._img = person_icon
        person_icon_lbl.pack(side="left", padx=(0, 4))
    tk.Label(person_label_frame, text="PERSON TO SEE (Optional)",
             font=("Segoe UI", 8, "bold"),
             bg=C_WHITE, fg=C_LABEL).pack(side="left")
    person_container = tk.Frame(person_frame, bg=C_INPUT_BG,
                                 highlightbackground=C_BORDER,
                                 highlightcolor=C_RED_ACCENT,
                                 highlightthickness=1)
    person_container.pack(fill="x")
    person_entry = tk.Entry(person_container, textvariable=person_var,
                             font=("Segoe UI", 10), bg=C_INPUT_BG, fg=C_TEXT_MID,
                             insertbackground=C_RED_ACCENT, relief="flat", bd=0)
    person_entry.pack(fill="x", ipady=9, padx=12)

    def _toggle_other(*_):
        if purpose_var.get() == "Other":
            if not other_row.winfo_ismapped():
                other_row.pack(fill="x", pady=(0, 16), before=row5)
            _animate_other(True)
        else:
            other_purpose_var.set("")
            _animate_other(False)

    purpose_var.trace_add("write", _toggle_other)

    # Focus effects
    for entry, container in [
        (name_entry, name_container),
        (contact_entry, contact_container),
        (addr_entry, addr_container),
        (date_entry, date_container),
        (time_entry, time_container),
        (person_entry, person_container),
        (other_entry, other_container),
        (purpose_combo, purpose_container),
    ]:
        def _fi(e, c=container): c.config(highlightbackground=C_RED_ACCENT, highlightthickness=2)
        def _fo(e, c=container): c.config(highlightbackground=C_BORDER, highlightthickness=1)
        entry.bind("<FocusIn>", _fi)
        entry.bind("<FocusOut>", _fo)

    # Submit button
    btn_frame = tk.Frame(card, bg=C_WHITE, padx=30)
    btn_frame.pack(fill="x", pady=(4, 0))

    submit_icon = load_icon_image("paper-plane.png", 18)
    submit_btn = RoundedButton(btn_frame, text="REGISTER VISIT",
                               icon=submit_icon, compound="left",
                               bg=C_RED_ACCENT,
                               activebackground=C_RED_DARK,
                               min_width=300)
    if submit_icon:
        submit_btn._icon_ref = submit_icon
    submit_btn.pack(pady=(8, 0), anchor="center")

    # Security notice
    notice_frame = tk.Frame(card, bg=C_WHITE)
    notice_frame.pack(fill="x", pady=(10, 24))
    tk.Label(notice_frame,
             text="🔒  Your information is secure and will be used for barangay records only.",
             font=("Segoe UI", 8, "italic"),
             bg=C_WHITE, fg=C_TEXT_MID).pack()

    def submit_visitor():
        name    = name_var.get().strip()
        contact = contact_var.get().strip()
        address = address_var.get().strip()
        purpose = purpose_var.get().strip()
        if purpose == "Other":
            other = other_purpose_var.get().strip()
            if other: purpose = other
        date   = date_var.get().strip()
        time   = time_var.get().strip()
        person = person_var.get().strip()

        errors = []
        if not name:    errors.append("• Full Name is required.")
        if not contact: errors.append("• Contact Number is required.")
        elif not is_valid_contact(contact): errors.append("• Contact must be 09XXXXXXXXX or +639XXXXXXXXX.")
        if not address: errors.append("• Address is required.")
        if not purpose or purpose == "Select purpose of visit":
            errors.append("• Please select a purpose of visit.")
        if purpose_var.get() == "Other" and not other_purpose_var.get().strip():
            errors.append("• Please specify the purpose of visit.")
        if not is_valid_date(date):  errors.append("• Date must be YYYY-MM-DD.")
        if not is_valid_time(time):  errors.append("• Time must be HH:MM.")

        if errors:
            messagebox.showerror("Incomplete Form", "\n".join(errors)); return

        try:
            insert_visitor(name, contact, address, purpose, date, time, person)
            messagebox.showinfo("Visit Registered! ✓",
                                f"Welcome, {name}!\n\nYour visit has been successfully logged.\nThank you for visiting Barangay San Andres.")
            name_var.set(""); contact_var.set(""); address_var.set("")
            purpose_var.set("Select purpose of visit"); other_purpose_var.set("")
            person_var.set("")
            other_row.pack_forget()
            date_var.set(datetime.now().strftime("%Y-%m-%d"))
            time_var.set(datetime.now().strftime("%I:%M %p"))
        except Exception as error:
            messagebox.showerror("Error", str(error))

    submit_btn.config(command=submit_visitor)


def try_admin_login(root, main_frame, sidebar_nav=None):
    if open_login_window(root):
        show_admin_screen(root, main_frame, sidebar_nav)


# ============================================================
#  ANALYTICS HELPERS
# ============================================================

def analytics_get_daily_counts(days=30):
    """Returns list of (date_str, count) for the last `days` days."""
    from datetime import timedelta
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    results = []
    today = datetime.now().date()
    for i in range(days - 1, -1, -1):
        d = (today - timedelta(days=i)).strftime("%Y-%m-%d")
        cur.execute("SELECT COUNT(*) FROM visitors WHERE visit_date=?", (d,))
        results.append((d, cur.fetchone()[0]))
    conn.close()
    return results

def analytics_get_weekly_counts(weeks=12):
    """Returns list of (week_label, count) for the last `weeks` weeks."""
    from datetime import timedelta
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    results = []
    today = datetime.now().date()
    for i in range(weeks - 1, -1, -1):
        start = today - timedelta(days=today.weekday()) - timedelta(weeks=i)
        end   = start + timedelta(days=6)
        cur.execute("SELECT COUNT(*) FROM visitors WHERE visit_date BETWEEN ? AND ?",
                    (start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")))
        label = f"Wk {start.strftime('%m/%d')}"
        results.append((label, cur.fetchone()[0]))
    conn.close()
    return results

def analytics_get_monthly_counts(months=12):
    """Returns list of (month_label, count) for the last `months` months."""
    from datetime import timedelta
    import calendar
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    results = []
    today = datetime.now()
    for i in range(months - 1, -1, -1):
        month = today.month - i
        year  = today.year
        while month <= 0:
            month += 12
            year  -= 1
        _, last_day = calendar.monthrange(year, month)
        start = f"{year}-{month:02d}-01"
        end   = f"{year}-{month:02d}-{last_day:02d}"
        cur.execute("SELECT COUNT(*) FROM visitors WHERE visit_date BETWEEN ? AND ?", (start, end))
        label = datetime(year, month, 1).strftime("%b %Y")
        results.append((label, cur.fetchone()[0]))
    conn.close()
    return results

def analytics_get_purpose_counts():
    """Returns dict of purpose -> count."""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    purposes = [
        "Business / Transaction",
        "Certificate Request",
        "Complaint / Concern",
        "Clearance Application",
        "Meeting / Appointment",
        "Social Service Inquiry",
        "Other",
    ]
    results = {}
    for p in purposes:
        cur.execute("SELECT COUNT(*) FROM visitors WHERE purpose=?", (p,))
        results[p] = cur.fetchone()[0]
    # Catch anything not in the standard list (e.g. free-text "Other" entries)
    formatted = ", ".join(f"'{x}'" for x in purposes)
    cur.execute(f"SELECT COUNT(*) FROM visitors WHERE purpose NOT IN ({formatted})")
    extra = cur.fetchone()[0]
    results["Other"] = results.get("Other", 0) + extra
    conn.close()
    return results

def analytics_get_hourly_counts():
    """Returns list of (hour_label, count) for hours 0-23."""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    results = []
    for h in range(24):
        hstr = f"{h:02d}:"
        cur.execute("SELECT COUNT(*) FROM visitors WHERE visit_time LIKE ?", (hstr + "%",))
        label = f"{h:02d}:00"
        results.append((label, cur.fetchone()[0]))
    conn.close()
    return results

def analytics_get_top_persons(limit=10):
    """Returns list of (person_to_see, count) top visited persons, case-insensitive grouping."""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        SELECT UPPER(TRIM(person_to_see)) as name, COUNT(*) as cnt FROM visitors
        WHERE person_to_see IS NOT NULL AND TRIM(person_to_see) != ''
        GROUP BY UPPER(TRIM(person_to_see)) ORDER BY cnt DESC LIMIT ?
    """, (limit,))
    results = [(name.title(), cnt) for name, cnt in cur.fetchall()]
    conn.close()
    return results


# ============================================================
#  ANALYTICS SCREEN  –  YouTube-style dark analytics (same window)
# ============================================================

def show_analytics_screen(root, main_frame, sidebar_nav=None):
    for w in main_frame.winfo_children():
        w.destroy()

    if sidebar_nav:
        sidebar_nav["update"]("visitor_log")

    # ── COLOR PALETTE (light mode) ──────────────────────────
    A_BG        = "#F5F5F5"
    A_SURFACE   = "#FFFFFF"
    A_SURFACE2  = "#F0F0F0"
    A_BORDER    = "#DEDEDE"
    A_TEXT      = "#1A1A1A"
    A_MUTED     = "#777777"
    A_BLUE      = "#1A73E8"
    A_RED       = "#CC0000"
    A_GREEN     = "#0A8A5C"
    A_GOLD      = "#C47D00"
    A_PURPLE    = "#7B2FBE"
    A_TEAL      = "#007B9E"

    BAR_COLORS = [A_BLUE, A_RED, A_GREEN, A_GOLD, A_PURPLE, A_TEAL,
                  "#E74C3C", "#2ECC71", "#F39C12", "#1ABC9C", "#E67E22", "#8E44AD"]

    main_frame.configure(bg=A_BG)

    # ── TOP HEADER ──────────────────────────────────────────
    header = tk.Frame(main_frame, bg=C_RED_DARK, height=56)
    header.pack(fill="x")
    header.pack_propagate(False)
    tk.Label(header, text="  📈  Analytics",
             font=("Segoe UI", 15, "bold"),
             bg=C_RED_DARK, fg=C_WHITE).pack(side="left", padx=20, fill="y")
    tk.Label(header, text="Visitor Insights — Barangay San Andres",
             font=("Segoe UI", 10),
             bg=C_RED_DARK, fg=C_TEXT_LIGHT).pack(side="left", fill="y")

    # "Visitor Log" button — goes back to admin
    back_to_log_btn = RoundedButton(header, text="📋  Visitor Log",
                                    bg=C_RED_MID,
                                    activebackground=C_RED_DARK,
                                    min_width=160,
                                    command=lambda: show_admin_screen(root, main_frame, sidebar_nav))
    back_to_log_btn.pack(side="right", padx=14, pady=8)

    tk.Frame(main_frame, bg=A_BLUE, height=2).pack(fill="x")

    # ── TAB BAR ─────────────────────────────────────────────
    tab_bar = tk.Frame(main_frame, bg=A_SURFACE, height=44)
    tab_bar.pack(fill="x")
    tab_bar.pack_propagate(False)

    active_tab = [0]
    tab_frames  = {}
    tab_buttons = {}

    def make_tab_btn(parent, label, idx):
        f = tk.Frame(parent, bg=A_SURFACE, cursor="hand2")
        f.pack(side="left")
        indicator = tk.Frame(f, bg=A_BLUE if idx == 0 else A_SURFACE, height=3)
        indicator.pack(fill="x", side="bottom")
        lbl = tk.Label(f, text=label,
                       font=("Segoe UI", 10, "bold" if idx == 0 else "normal"),
                       bg=A_SURFACE, fg=A_TEXT if idx == 0 else A_MUTED,
                       padx=20, pady=12, cursor="hand2")
        lbl.pack()
        tab_buttons[idx] = (lbl, indicator)

        def on_click(e, i=idx):
            if active_tab[0] == i:
                return
            if i not in tab_frames:
                return
            old = active_tab[0]
            ob, oi = tab_buttons[old]
            ob.config(fg=A_MUTED, font=("Segoe UI", 10))
            oi.config(bg=A_SURFACE)
            active_tab[0] = i
            lbl.config(fg=A_TEXT, font=("Segoe UI", 10, "bold"))
            indicator.config(bg=A_BLUE)
            for k, fr in tab_frames.items():
                fr.pack_forget()
            tab_frames[i].pack(fill="both", expand=True)

        for w in [f, lbl]:
            w.bind("<Button-1>", on_click)

    make_tab_btn(tab_bar, "Overview",      0)
    make_tab_btn(tab_bar, "Purpose",       1)
    make_tab_btn(tab_bar, "Time of Visit", 2)
    make_tab_btn(tab_bar, "Person to See", 3)

    tk.Frame(tab_bar, bg=A_BORDER, height=1).pack(side="bottom", fill="x")

    # ── SCROLL CONTAINER HELPER ──────────────────────────────
    def make_scroll_frame(parent):
        outer = tk.Frame(parent, bg=A_BG)
        canvas = tk.Canvas(outer, bg=A_BG, highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=A_BG)
        win_id = canvas.create_window(0, 0, anchor="nw", window=inner)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))

        def _on_mw(e):
            try:
                delta = int(-1*(e.delta/120))
                # Clamp: don't scroll above top (0.0) or below bottom (1.0)
                top, bottom = canvas.yview()
                if delta < 0 and top <= 0.0:
                    return
                if delta > 0 and bottom >= 1.0:
                    return
                canvas.yview_scroll(delta, "units")
            except Exception:
                pass
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mw))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        return outer, inner

    def ana_card(parent, title, subtitle=""):
        card = tk.Frame(parent, bg=A_SURFACE,
                        highlightbackground=A_BORDER, highlightthickness=1)
        card.pack(fill="x", padx=24, pady=(16, 0))
        tk.Frame(card, bg=A_BLUE, height=3).pack(fill="x")
        body = tk.Frame(card, bg=A_SURFACE, padx=24, pady=18)
        body.pack(fill="both", expand=True)
        hrow = tk.Frame(body, bg=A_SURFACE)
        hrow.pack(fill="x", pady=(0, 4))
        tk.Label(hrow, text=title, font=("Segoe UI", 12, "bold"),
                 bg=A_SURFACE, fg=A_TEXT).pack(side="left")
        if subtitle:
            tk.Label(hrow, text=f"  {subtitle}", font=("Segoe UI", 9),
                     bg=A_SURFACE, fg=A_MUTED).pack(side="left")
        tk.Frame(body, bg=A_BORDER, height=1).pack(fill="x", pady=(8, 14))
        return body

    def draw_bar_chart(parent, data, bar_color=A_BLUE, height=220, bg=A_SURFACE,
                       show_values=True, label_every=1, rotate_labels=False, initial_zoom=1.0):
        if not data:
            tk.Label(parent, text="No data available.", font=("Segoe UI", 10),
                     bg=bg, fg=A_MUTED).pack(pady=20)
            return
        labels = [d[0] for d in data]
        values = [d[1] for d in data]
        max_val = max(values) if max(values) > 0 else 1

        # zoom / scroll state
        state = {"zoom": initial_zoom, "scroll_x": 0}
        MIN_ZOOM, MAX_ZOOM = 1.0, 8.0

        wrapper = tk.Frame(parent, bg=bg)
        wrapper.pack(fill="x", pady=(0, 4))

        c = tk.Canvas(wrapper, bg=bg, height=height, highlightthickness=0)
        c.pack(fill="x", expand=True)

        hbar = tk.Scrollbar(wrapper, orient="horizontal")

        margin_l, margin_r, margin_top, margin_bot = 62, 16, 16, 38

        def _redraw(e=None):
            c.delete("all")
            view_w = c.winfo_width()
            if view_w < 50:
                return
            chart_h = height - margin_top - margin_bot
            n = len(labels)
            zoom = state["zoom"]
            base_bar_w = max(8, min(60, (view_w - margin_l - margin_r) // max(n, 1) - 4))
            bar_w = int(base_bar_w * zoom)
            gap = max(2, int(4 * zoom))
            total_content_w = n * (bar_w + gap) + gap
            max_scroll = max(0, total_content_w - (view_w - margin_l - margin_r))
            state["scroll_x"] = max(0, min(state["scroll_x"], max_scroll))
            off = state["scroll_x"]

            if zoom > 1.0 and max_scroll > 0:
                hbar.pack(fill="x")
                lo = off / total_content_w if total_content_w > 0 else 0
                hi = lo + (view_w - margin_l - margin_r) / total_content_w
                hbar.set(lo, min(hi, 1.0))
            else:
                hbar.pack_forget()

            # Y-axis grid + labels (draw first, will be covered by mask, then redrawn)
            for step in range(0, 6):
                y = margin_top + chart_h - int(chart_h * step / 5)
                c.create_line(margin_l, y, view_w, y, fill=A_BORDER, dash=(4, 4))

            # auto-thin labels based on bar width
            show_every = max(1, int(20 / max(bar_w, 1)))

            for i, (lbl, val) in enumerate(zip(labels, values)):
                cx = margin_l + gap + i * (bar_w + gap) - off
                x1, x2 = cx, cx + bar_w
                if x2 < margin_l or x1 > view_w:
                    continue
                bar_h = int(chart_h * val / max_val) if max_val > 0 else 0
                y1 = margin_top + chart_h - bar_h
                y2 = margin_top + chart_h
                color = bar_color if isinstance(bar_color, str) else bar_color[i % len(bar_color)]
                draw_x1 = max(x1, margin_l)
                if draw_x1 < x2:
                    c.create_rectangle(draw_x1, y1, x2, y2, fill=color, outline="", width=0)
                if show_values and val > 0 and x1 >= margin_l:
                    c.create_text((x1 + x2) // 2, y1 - 5, text=str(val), anchor="s",
                                   fill="#222222", font=("Segoe UI", 8, "bold"))
                mid_x = (x1 + x2) // 2
                if i % show_every == 0 and mid_x > margin_l:
                    c.create_text(mid_x, y2 + 5, text=lbl,
                                   anchor="n", fill="#333333", font=("Segoe UI", 8, "bold"))

            # Left mask to hide bars sliding under Y-axis
            c.create_rectangle(0, 0, margin_l - 1, height + 4, fill=bg, outline="")
            # Redraw Y labels on top of mask
            for step in range(0, 6):
                y = margin_top + chart_h - int(chart_h * step / 5)
                val_label = int(max_val * step / 5)
                c.create_text(margin_l - 6, y, text=str(val_label), anchor="e",
                               fill="#444444", font=("Segoe UI", 8, "bold"))
            # Axes
            c.create_line(margin_l, margin_top, margin_l, margin_top + chart_h,
                          fill=A_MUTED, width=1)
            c.create_line(margin_l, margin_top + chart_h, view_w, margin_top + chart_h,
                          fill=A_MUTED, width=1)

        def _on_hbar_move(cmd, *args):
            view_w = c.winfo_width()
            n = len(labels)
            zoom = state["zoom"]
            base_bar_w = max(8, min(60, (view_w - margin_l - margin_r) // max(n, 1) - 4))
            bar_w = int(base_bar_w * zoom)
            gap = max(2, int(4 * zoom))
            total_content_w = n * (bar_w + gap) + gap
            if cmd == "moveto":
                state["scroll_x"] = float(args[0]) * total_content_w
            elif cmd == "scroll":
                state["scroll_x"] += int(args[0]) * 30
            _redraw()

        hbar.config(command=_on_hbar_move)

        def _on_wheel(event):
            if event.state & 0x4:  # Ctrl held = zoom
                factor = 1.15 if event.delta > 0 else (1 / 1.15)
                new_zoom = max(MIN_ZOOM, min(MAX_ZOOM, state["zoom"] * factor))
                cursor_chart_x = event.x - margin_l + state["scroll_x"]
                state["zoom"] = new_zoom
                state["scroll_x"] = max(0, cursor_chart_x - (event.x - margin_l))
                _redraw()
                return "break"
            else:  # plain scroll = horizontal pan when zoomed
                if state["zoom"] > 1.0:
                    state["scroll_x"] += int(-event.delta / 120 * 40)
                    _redraw()
                    return "break"

        tk.Label(wrapper, text="Ctrl+Scroll to zoom  •  Scroll to pan horizontally",
                 font=("Segoe UI", 7), bg=bg, fg=A_MUTED).pack(anchor="e", padx=4)

        _init_done = [False]

        def _on_configure(e=None):
            _redraw()

        def _initial_draw():
            if initial_zoom > 1.0:
                view_w = c.winfo_width()
                if view_w < 50:
                    c.after(100, _initial_draw)
                    return
                n = len(labels)
                zoom = state["zoom"]
                base_bar_w = max(8, min(60, (view_w - margin_l - margin_r) // max(n, 1) - 4))
                bar_w = int(base_bar_w * zoom)
                gap = max(2, int(4 * zoom))
                total_content_w = n * (bar_w + gap) + gap
                state["scroll_x"] = max(0, total_content_w - (view_w - margin_l - margin_r))
            _redraw()
            _init_done[0] = True

        c.bind("<Configure>", _on_configure)
        c.bind("<MouseWheel>", _on_wheel)
        c.after(200, _initial_draw)

    def stat_pill(parent, label, value, color):
        f = tk.Frame(parent, bg=A_SURFACE2,
                     highlightbackground=color, highlightthickness=2)
        f.pack(side="left", padx=8, pady=4)
        tk.Label(f, text=value, font=("Segoe UI", 22, "bold"),
                 bg=A_SURFACE2, fg=color).pack(padx=20, pady=(12, 2))
        tk.Label(f, text=label, font=("Segoe UI", 9),
                 bg=A_SURFACE2, fg=A_MUTED).pack(padx=20, pady=(0, 12))

    # ══════════════════════════════════════════════════════════
    #  TAB 0 — OVERVIEW
    # ══════════════════════════════════════════════════════════
    overview_outer, overview = make_scroll_frame(main_frame)
    tab_frames[0] = overview_outer

    def build_overview(parent):
        from datetime import timedelta
        today_str   = datetime.now().strftime("%Y-%m-%d")
        week_start  = (datetime.now().date() - timedelta(days=datetime.now().weekday())).strftime("%Y-%m-%d")
        month_start = datetime.now().strftime("%Y-%m-01")

        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM visitors WHERE visit_date=?", (today_str,))
        today_cnt = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM visitors WHERE visit_date >= ?", (week_start,))
        week_cnt = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM visitors WHERE visit_date >= ?", (month_start,))
        month_cnt = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM visitors")
        total_cnt = cur.fetchone()[0]
        conn.close()

        pill_card = ana_card(parent, "📊 Summary",
                              f"as of {datetime.now().strftime('%B %d, %Y %I:%M %p')}")
        pills_row = tk.Frame(pill_card, bg=A_SURFACE)
        pills_row.pack(fill="x", pady=4)
        stat_pill(pills_row, "Today",      str(today_cnt), A_BLUE)
        stat_pill(pills_row, "This Week",  str(week_cnt),  A_GREEN)
        stat_pill(pills_row, "This Month", str(month_cnt), A_GOLD)
        stat_pill(pills_row, "All Time",   str(total_cnt), A_RED)

        daily_card = ana_card(parent, "📅 Daily Visitors", "Last 30 days")
        daily_data = analytics_get_daily_counts(30)
        draw_bar_chart(daily_card, daily_data, bar_color=A_BLUE, height=240,
                       bg=A_SURFACE, label_every=max(1, len(daily_data)//10), rotate_labels=True,
                       initial_zoom=3.0)

        week_card = ana_card(parent, "📆 Weekly Visitors", "Last 12 weeks")
        draw_bar_chart(week_card, analytics_get_weekly_counts(12), bar_color=A_GREEN,
                       height=220, bg=A_SURFACE)

        month_card = ana_card(parent, "🗓 Monthly Visitors", "Last 12 months")
        draw_bar_chart(month_card, analytics_get_monthly_counts(12), bar_color=A_GOLD,
                       height=220, bg=A_SURFACE)

        tk.Frame(parent, bg=A_BG, height=30).pack()

    build_overview(overview)

    # ══════════════════════════════════════════════════════════
    #  TAB 1 — PURPOSE
    # ══════════════════════════════════════════════════════════
    purpose_outer, purpose_frame = make_scroll_frame(main_frame)
    tab_frames[1] = purpose_outer

    def build_purpose(parent):
        purpose_data = analytics_get_purpose_counts()
        total = sum(purpose_data.values()) or 1
        short_labels = {
            "Business / Transaction": "Business",
            "Certificate Request":    "Certificate",
            "Complaint / Concern":    "Complaint",
            "Clearance Application":  "Clearance",
            "Meeting / Appointment":  "Meeting",
            "Social Service Inquiry": "Social Svc",
            "Other":                  "Other",
        }
        sorted_purposes = sorted(purpose_data.items(), key=lambda x: x[1], reverse=True)

        bar_card = ana_card(parent, "📋 Visits by Purpose", "All time breakdown")
        bar_data = [(short_labels.get(p, p), v) for p, v in sorted_purposes]
        draw_bar_chart(bar_card, bar_data, bar_color=BAR_COLORS, height=260, bg=A_SURFACE)

        tbl_card = ana_card(parent, "📊 Purpose Details")
        for i, (purpose, count) in enumerate(sorted_purposes):
            row_bg = A_SURFACE if i % 2 == 0 else A_SURFACE2
            row = tk.Frame(tbl_card, bg=row_bg)
            row.pack(fill="x", pady=1)
            color = BAR_COLORS[i % len(BAR_COLORS)]
            tk.Label(row, text="█", font=("Segoe UI", 11), bg=row_bg, fg=color).pack(
                side="left", padx=(8, 4), pady=6)
            tk.Label(row, text=purpose, font=("Segoe UI", 10), bg=row_bg, fg=A_TEXT,
                     anchor="w").pack(side="left", fill="x", expand=True, pady=6)
            pct = f"{count / total * 100:.1f}%"
            tk.Label(row, text=pct, font=("Segoe UI", 9), bg=row_bg, fg=A_MUTED).pack(
                side="right", padx=8)
            tk.Label(row, text=str(count), font=("Segoe UI", 10, "bold"),
                     bg=row_bg, fg=color).pack(side="right", padx=16)
            prog_outer = tk.Frame(tbl_card, bg=A_SURFACE2, height=4)
            prog_outer.pack(fill="x", padx=8, pady=(0, 2))
            prog_outer.pack_propagate(False)
            tk.Frame(prog_outer, bg=color, height=4).place(relwidth=count/total, relheight=1.0)

        tk.Frame(parent, bg=A_BG, height=30).pack()

    build_purpose(purpose_frame)

    # ══════════════════════════════════════════════════════════
    #  TAB 2 — TIME
    # ══════════════════════════════════════════════════════════
    time_outer, time_frame = make_scroll_frame(main_frame)
    tab_frames[2] = time_outer

    def build_time(parent):
        hourly_data = analytics_get_hourly_counts()
        peak_hour   = max(hourly_data, key=lambda x: x[1], default=("--", 0))

        hour_card = ana_card(parent, "⏰ Hourly Visit Distribution", "All records — 24-hour view")
        peak_row = tk.Frame(hour_card, bg=A_SURFACE)
        peak_row.pack(fill="x", pady=(0, 12))
        tk.Label(peak_row, text="🔥  Peak Hour: ", font=("Segoe UI", 10, "bold"),
                 bg=A_SURFACE, fg=A_MUTED).pack(side="left")
        tk.Label(peak_row, text=f"{peak_hour[0]}  ({peak_hour[1]} visitors)",
                 font=("Segoe UI", 11, "bold"), bg=A_SURFACE, fg=A_GOLD).pack(side="left")
        draw_bar_chart(hour_card, hourly_data, bar_color=A_TEAL, height=280,
                       bg=A_SURFACE, label_every=2)

        am_total = sum(cnt for lbl, cnt in hourly_data[:12])
        pm_total = sum(cnt for lbl, cnt in hourly_data[12:])
        split_card = ana_card(parent, "🌤 AM vs PM Breakdown")
        split_row = tk.Frame(split_card, bg=A_SURFACE)
        split_row.pack(fill="x")
        stat_pill(split_row, "Morning (AM)",   str(am_total), A_GOLD)
        stat_pill(split_row, "Afternoon (PM)", str(pm_total), A_PURPLE)

        tk.Frame(parent, bg=A_BG, height=30).pack()

    build_time(time_frame)

    # ══════════════════════════════════════════════════════════
    #  TAB 3 — PERSON TO SEE
    # ══════════════════════════════════════════════════════════
    person_outer, person_frame = make_scroll_frame(main_frame)
    tab_frames[3] = person_outer

    def build_person(parent):
        persons = analytics_get_top_persons(15)
        person_card = ana_card(parent, "👤 Top Persons Visited", "Most frequently requested")

        if not persons:
            tk.Label(person_card, text="No person-to-see data recorded yet.",
                     font=("Segoe UI", 10), bg=A_SURFACE, fg=A_MUTED).pack(pady=20)
        else:
            max_cnt = persons[0][1] if persons else 1
            rank_color = [A_GOLD, "#C0C0C0", "#CD7F32"] + [A_MUTED] * 20
            for i, (name, cnt) in enumerate(persons):
                row_bg = A_SURFACE if i % 2 == 0 else A_SURFACE2
                row = tk.Frame(person_card, bg=row_bg)
                row.pack(fill="x", pady=2)
                tk.Label(row, text=f"#{i+1}", font=("Segoe UI", 10, "bold"),
                         bg=row_bg, fg=rank_color[i], width=4).pack(side="left", pady=8)
                tk.Label(row, text=name, font=("Segoe UI", 10), bg=row_bg,
                         fg=A_TEXT, anchor="w").pack(side="left", fill="x", expand=True)
                tk.Label(row, text=f"{cnt} visit{'s' if cnt != 1 else ''}",
                         font=("Segoe UI", 10, "bold"), bg=row_bg, fg=A_BLUE).pack(
                    side="right", padx=16)
                prog_outer = tk.Frame(person_card, bg=A_SURFACE2, height=5)
                prog_outer.pack(fill="x", padx=8, pady=(0, 2))
                prog_outer.pack_propagate(False)
                tk.Frame(prog_outer, bg=A_BLUE, height=5).place(
                    relwidth=cnt / max_cnt, relheight=1.0)

        if persons:
            chart_card = ana_card(parent, "📊 Visit Count Chart")
            chart_data = [(p[:18] + "…" if len(p) > 18 else p, c) for p, c in persons[:10]]
            draw_bar_chart(chart_card, chart_data, bar_color=BAR_COLORS, height=260,
                           bg=A_SURFACE, label_every=1, rotate_labels=True)

        tk.Frame(parent, bg=A_BG, height=30).pack()

    build_person(person_frame)

    # Show first tab
    tab_frames[0].pack(fill="both", expand=True)



def show_admin_screen(root, main_frame, sidebar_nav=None):
    for w in main_frame.winfo_children():
        w.destroy()

    # Inactivity auto-logout (5 minutes)
    _inactivity_after_id = [None]
    INACTIVITY_SECONDS = 5 * 60

    _inactivity_bind_ids = [None, None, None]

    def _cancel_inactivity_timer():
        try:
            if _inactivity_after_id[0]:
                root.after_cancel(_inactivity_after_id[0])
                _inactivity_after_id[0] = None
        except Exception:
            _inactivity_after_id[0] = None
        # Only unbind the specific bind IDs we registered
        for i, event in enumerate(['<KeyPress>', '<Motion>', '<ButtonPress>']):
            try:
                if _inactivity_bind_ids[i]:
                    root.unbind(event, _inactivity_bind_ids[i])
                    _inactivity_bind_ids[i] = None
            except Exception:
                _inactivity_bind_ids[i] = None

    def _do_inactivity_logout():
        _cancel_inactivity_timer()
        try:
            messagebox.showinfo("Logged Out", "Admin logged out due to inactivity.", parent=root)
        except Exception:
            pass
        show_user_screen(root, main_frame, sidebar_nav)

    def _reset_inactivity_timer(event=None):
        try:
            if _inactivity_after_id[0]:
                root.after_cancel(_inactivity_after_id[0])
        except Exception:
            pass
        _inactivity_after_id[0] = root.after(INACTIVITY_SECONDS * 1000, _do_inactivity_logout)

    def _bind_inactivity_events():
        # Use root.bind (not bind_all) so bindings don't bleed into other screens
        _inactivity_bind_ids[0] = root.bind('<KeyPress>',    _reset_inactivity_timer, '+')
        _inactivity_bind_ids[1] = root.bind('<Motion>',      _reset_inactivity_timer, '+')
        _inactivity_bind_ids[2] = root.bind('<ButtonPress>', _reset_inactivity_timer, '+')

    # Start listening for activity
    _bind_inactivity_events()
    _reset_inactivity_timer()

    if sidebar_nav:
        sidebar_nav["update"]("visitor_log")

    # ── TOP BAR ──────────────────────────────────────────────
    topbar = tk.Frame(main_frame, bg=C_RED_DARK, height=56)
    topbar.pack(fill="x")
    topbar.pack_propagate(False)

    back_icon = load_icon_image_file("back.png", 16)
    def _back_to_user():
        _cancel_inactivity_timer()
        show_user_screen(root, main_frame, sidebar_nav)

    back_btn = RoundedButton(topbar, text="Back to Visitor Entry",
                             icon=back_icon, compound="left",
                             bg=C_RED_MID,
                             activebackground=C_RED_ACCENT,
                             min_width=260,
                             command=_back_to_user)
    if back_icon:
        back_btn._icon_ref = back_icon
    back_btn.pack(side="right", padx=14, pady=8)

    tk.Label(topbar, text="  🛡  ADMIN PANEL  —  Visitor Records",
             font=("Georgia", 12, "bold"),
             bg=C_RED_DARK, fg=C_WHITE).pack(side="left", padx=16, fill="y")

    tk.Frame(main_frame, bg=C_GOLD, height=3).pack(fill="x")

    # ── SEARCH + ACTIONS BAR ─────────────────────────────────
    toolbar = tk.Frame(main_frame, bg=C_OFF_WHITE, pady=10)
    toolbar.pack(fill="x", padx=16)

    # Search
    search_group = tk.Frame(toolbar, bg=C_OFF_WHITE)
    search_group.pack(side="left")

    tk.Label(search_group, text="🔍", font=("Segoe UI", 11),
             bg=C_OFF_WHITE, fg=C_RED_ACCENT).pack(side="left", padx=(0, 6))

    search_var = tk.StringVar()
    search_container = tk.Frame(search_group, bg=C_WHITE,
                                 highlightbackground=C_BORDER,
                                 highlightcolor=C_RED_ACCENT,
                                 highlightthickness=1)
    search_container.pack(side="left")
    search_entry = tk.Entry(search_container, textvariable=search_var,
                             font=("Segoe UI", 10), width=28,
                             bg=C_WHITE, fg=C_TEXT_DARK,
                             relief="flat", bd=0,
                             insertbackground=C_RED_ACCENT)
    search_entry.pack(ipady=7, padx=10)
    search_entry.bind("<FocusIn>",
                      lambda e: search_container.config(highlightbackground=C_RED_ACCENT, highlightthickness=2))
    search_entry.bind("<FocusOut>",
                      lambda e: search_container.config(highlightbackground=C_BORDER, highlightthickness=1))

    record_count_var = tk.StringVar(value="")
    tk.Label(search_group, textvariable=record_count_var,
             font=("Segoe UI", 8), bg=C_OFF_WHITE, fg=C_TEXT_MID).pack(side="left", padx=10)

    # Action buttons right side
    action_group = tk.Frame(toolbar, bg=C_OFF_WHITE)
    action_group.pack(side="right")

    # Date range for export
    date_group = tk.Frame(action_group, bg=C_OFF_WHITE)
    date_group.pack(side="left", padx=(12, 0))

    tk.Label(date_group, text="From:", font=("Segoe UI", 8, "bold"),
             bg=C_OFF_WHITE, fg=C_TEXT_DARK).pack(side="left", padx=(0, 4))
    start_date_var = tk.StringVar()
    s_cont = tk.Frame(date_group, bg=C_WHITE,
                      highlightbackground=C_BORDER, highlightthickness=1)
    s_cont.pack(side="left", padx=(0, 10))
    tk.Entry(s_cont, textvariable=start_date_var, font=("Segoe UI", 9),
             width=11, bg=C_WHITE, fg=C_TEXT_DARK,
             relief="flat", bd=0).pack(ipady=5, padx=8)

    tk.Label(date_group, text="To:", font=("Segoe UI", 8, "bold"),
             bg=C_OFF_WHITE, fg=C_TEXT_DARK).pack(side="left", padx=(0, 4))
    end_date_var = tk.StringVar()
    e_cont = tk.Frame(date_group, bg=C_WHITE,
                      highlightbackground=C_BORDER, highlightthickness=1)
    e_cont.pack(side="left", padx=(0, 10))
    tk.Entry(e_cont, textvariable=end_date_var, font=("Segoe UI", 9),
             width=11, bg=C_WHITE, fg=C_TEXT_DARK,
             relief="flat", bd=0).pack(ipady=5, padx=8)

    def _run_query(query, params=()):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        conn.close()
        return rows

    def export_to_excel(filename, rows, date_range):
        """Export records to Excel file"""
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl library not installed. Please install it with: pip install openpyxl")
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Visitor Records"
            
            # Style definitions
            header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
            header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add title - Barangay name first
            ws['A1'] = "BARANGAY SAN ANDRES"
            ws['A1'].font = Font(name="Arial", bold=True, size=15)
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Add records title
            ws['A2'] = f"Visitor Records ({date_range})"
            ws['A2'].font = Font(name="Arial", bold=True, size=14)
            ws.merge_cells('A2:H2')
            ws['A2'].alignment = Alignment(horizontal="center")
            
            # Add export date
            ws['A3'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws.merge_cells('A3:H3')
            ws['A3'].font = Font(name="Arial", italic=True, size=11)
            ws['A3'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            
            # Add headers
            headers = ['ID', 'Full Name', 'Contact', 'Address', 'Purpose', 'Person to See', 'Date', 'Time']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Add data rows
            for row_idx, row_data in enumerate(rows, start=6):
                id_, name, contact, address, purpose, person_to_see, vdate, vtime = row_data
                row_values = [id_, name, contact, address, purpose, person_to_see or "", vdate, vtime]
                for col_idx, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.font = Font(name="Arial", size=12)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = border
            
            # Adjust column widths for printing
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 14
            ws.column_dimensions['G'].width = 13
            ws.column_dimensions['H'].width = 12
            
            ws.row_dimensions[1].height = 20
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 16
            ws.row_dimensions[5].height = 25
            
            # Set print options to fit all content on page
            ws.print_options.horizontalCentered = False
            ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5
            
            # Fit all columns to one page width when printing
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            
            # Set print area to include all data
            max_row = 5 + len(rows)
            ws.print_area = f'A1:H{max_row}'
            
            wb.save(filename)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to Excel: {str(e)}")
            return False

    def export_to_pdf(filename, rows, date_range):
        """Export records to PDF file"""
        if not REPORTLAB_AVAILABLE:
            messagebox.showerror("Error", "reportlab library not installed. Please install it with: pip install reportlab")
            return False
        
        try:
            from reportlab.lib.pagesizes import landscape, A4
            doc = SimpleDocTemplate(filename, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = []
            
            # Barangay name
            styles = getSampleStyleSheet()
            barangay_style = ParagraphStyle(
                'BarangayTitle',
                parent=styles['Heading1'],
                fontSize=14,
                textColor=colors.HexColor("#8B0000"),
                spaceAfter=3,
                alignment=1,
                fontName='Times-Roman'
            )
            story.append(Paragraph("BARANGAY SAN ANDRES", barangay_style))
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=12,
                textColor=colors.HexColor("#8B0000"),
                spaceAfter=6,
                alignment=1,
                fontName='Times-Roman'
            )
            story.append(Paragraph(f"Visitor Records ({date_range})", title_style))
            
            # Export date
            export_style = ParagraphStyle(
                'ExportDate',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.grey,
                spaceAfter=12,
                alignment=1,
                fontName='Times-Roman'
            )
            story.append(Paragraph(f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", export_style))
            
            # Prepare table data with wrapping paragraphs for long cell content
            header_cell_style = ParagraphStyle(
                'HeaderCellStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.whitesmoke,
                alignment=0,
                fontName='Times-Bold',
                spaceBefore=0,
                spaceAfter=0,
                leading=12,
                wordWrap='CJK'
            )
            data_cell_style = ParagraphStyle(
                'DataCellStyle',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.black,
                alignment=0,
                fontName='Times-Roman',
                spaceBefore=0,
                spaceAfter=0,
                leading=10,
                wordWrap='CJK'
            )
            
            table_data = [[Paragraph(h, header_cell_style) for h in ['ID', 'Full Name', 'Contact', 'Address', 'Purpose', 'Person to See', 'Date', 'Time']]]
            for row in rows:
                id_, name, contact, address, purpose, person_to_see, vdate, vtime = row
                row_data = [
                    Paragraph(str(id_), data_cell_style),
                    Paragraph(str(name), data_cell_style),
                    Paragraph(str(contact), data_cell_style),
                    Paragraph(str(address), data_cell_style),
                    Paragraph(str(purpose), data_cell_style),
                    Paragraph(str(person_to_see or ""), data_cell_style),
                    Paragraph(str(vdate), data_cell_style),
                    Paragraph(str(vtime), data_cell_style),
                ]
                table_data.append(row_data)
            
            table = Table(table_data, colWidths=[0.35*inch, 1.3*inch, 0.9*inch, 1.8*inch, 1.4*inch, 1.4*inch, 0.9*inch, 0.7*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#8B0000")),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('VALIGN', (0, 1), (-1, -1), 'TOP'),
                ('WORDWRAP', (0, 0), (-1, -1), 'CJK'),
                ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFE8E8")]),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ]))
            
            story.append(table)
            doc.build(story)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to PDF: {str(e)}")
            return False

    def export_records():
        s = start_date_var.get().strip()
        e = end_date_var.get().strip()
        if s and not is_valid_date(s):
            messagebox.showerror("Error", "Start date must be YYYY-MM-DD."); return
        if e and not is_valid_date(e):
            messagebox.showerror("Error", "End date must be YYYY-MM-DD."); return

        if s and e:
            q = "SELECT id, full_name, contact, address, purpose, person_to_see, visit_date, visit_time FROM visitors WHERE visit_date BETWEEN ? AND ? ORDER BY visit_date, visit_time"
            p, rng = (s, e), f"{s} to {e}"
        elif s:
            q = "SELECT id, full_name, contact, address, purpose, person_to_see, visit_date, visit_time FROM visitors WHERE visit_date >= ? ORDER BY visit_date, visit_time"
            p, rng = (s,), f"from {s}"
        elif e:
            q = "SELECT id, full_name, contact, address, purpose, person_to_see, visit_date, visit_time FROM visitors WHERE visit_date <= ? ORDER BY visit_date, visit_time"
            p, rng = (e,), f"up to {e}"
        else:
            q = "SELECT id, full_name, contact, address, purpose, person_to_see, visit_date, visit_time FROM visitors ORDER BY visit_date, visit_time"
            p, rng = (), "all records"

        try:
            rows = _run_query(q, p)
        except Exception as err:
            messagebox.showerror("Error", str(err)); return

        if not rows:
            messagebox.showinfo("No Records", f"No records found for {rng}."); return

        # Show format selection dialog
        format_window = tk.Toplevel(root)
        format_window.title("Export Format")
        format_window.geometry("380x320")
        format_window.resizable(False, False)
        format_window.transient(root)
        format_window.grab_set()
        format_window.configure(bg=C_OFF_WHITE)
        
        # Center window
        format_window.update_idletasks()
        x = root.winfo_x() + (root.winfo_width() - format_window.winfo_width()) // 2
        y = root.winfo_y() + (root.winfo_height() - format_window.winfo_height()) // 2
        format_window.geometry(f"+{x}+{y}")
        
        # Main frame
        main_frame = tk.Frame(format_window, bg=C_OFF_WHITE)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        tk.Label(main_frame, text="Choose Export Format", 
                 font=("Segoe UI", 14, "bold"), 
                 bg=C_OFF_WHITE, fg=C_TEXT_DARK).pack(pady=(0, 15))
        
        # Format selection
        selected_format = tk.StringVar(value="xlsx")
        
        # Excel button
        excel_frame = tk.Frame(main_frame, bg=C_WHITE, highlightbackground=C_BORDER, 
                               highlightthickness=1, cursor="hand2")
        excel_frame.pack(fill="x", pady=5)
        
        def select_excel():
            selected_format.set("xlsx")
            excel_frame.config(highlightbackground=C_RED_DARK, highlightthickness=2)
            pdf_frame.config(highlightbackground=C_BORDER, highlightthickness=1)
        
        # Excel icon (use xls.png)
        excel_icon = load_icon_image_file("xls.png", 18)
        rb_text = "  Excel (.xlsx)"
        if excel_icon:
            tk.Radiobutton(excel_frame, text=rb_text, image=excel_icon, compound="left",
                           variable=selected_format, value="xlsx", font=("Segoe UI", 11),
                           bg=C_WHITE, fg=C_TEXT_DARK, selectcolor=C_RED_LIGHT,
                           command=select_excel).pack(anchor="w", padx=10, pady=10)
            excel_frame._icon_ref = excel_icon
        else:
            tk.Radiobutton(excel_frame, text="📊 Excel (.xlsx)", variable=selected_format,
                           value="xlsx", font=("Segoe UI", 11), bg=C_WHITE,
                           fg=C_TEXT_DARK, selectcolor=C_RED_LIGHT,
                           command=select_excel).pack(anchor="w", padx=10, pady=10)
        
        # PDF button
        pdf_frame = tk.Frame(main_frame, bg=C_WHITE, highlightbackground=C_BORDER, 
                             highlightthickness=1, cursor="hand2")
        pdf_frame.pack(fill="x", pady=5)
        
        def select_pdf():
            selected_format.set("pdf")
            pdf_frame.config(highlightbackground=C_RED_DARK, highlightthickness=2)
            excel_frame.config(highlightbackground=C_BORDER, highlightthickness=1)
        
        # PDF icon (use pdf.png)
        pdf_icon = load_icon_image_file("pdf.png", 18)
        rb_text_pdf = "  PDF (.pdf)"
        if pdf_icon:
            tk.Radiobutton(pdf_frame, text=rb_text_pdf, image=pdf_icon, compound="left",
                           variable=selected_format, value="pdf", font=("Segoe UI", 11),
                           bg=C_WHITE, fg=C_TEXT_DARK, selectcolor=C_RED_LIGHT,
                           command=select_pdf).pack(anchor="w", padx=10, pady=10)
            pdf_frame._icon_ref = pdf_icon
        else:
            tk.Radiobutton(pdf_frame, text="📄 PDF (.pdf)", variable=selected_format,
                           value="pdf", font=("Segoe UI", 11), bg=C_WHITE,
                           fg=C_TEXT_DARK, selectcolor=C_RED_LIGHT,
                           command=select_pdf).pack(anchor="w", padx=10, pady=10)
        
        # Set Excel as default selection
        select_excel()
        
        # Buttons frame (center buttons)
        button_frame = tk.Frame(main_frame, bg=C_OFF_WHITE)
        button_frame.pack(fill="x", pady=(20, 0))

        # center_frame holds the buttons and is packed centered horizontally
        center_frame = tk.Frame(button_frame, bg=C_OFF_WHITE)
        center_frame.pack(anchor='center')
        
        def proceed_export():
            fmt = selected_format.get()
            format_window.destroy()
            
            # File extension and type mapping
            ext_map = {"xlsx": ".xlsx", "pdf": ".pdf"}
            type_map = {
                "xlsx": [("Excel files", "*.xlsx")],
                "pdf": [("PDF files", "*.pdf")]
            }
            
            fn = filedialog.asksaveasfilename(parent=root, 
                                              defaultextension=ext_map[fmt],
                                              filetypes=type_map[fmt],
                                              title="Save records as...")
            if not fn: return
            
            success = False
            if fmt == "xlsx":
                success = export_to_excel(fn, rows, rng)
            elif fmt == "pdf":
                success = export_to_pdf(fn, rows, rng)
            
            if success:
                messagebox.showinfo("Saved", f"Records saved to:\n{fn}")
        
        export_btn = RoundedButton(center_frame, text="✓  Export", bg=C_RED_DARK,
                       activebackground=C_RED_MID, min_width=100,
                       command=proceed_export)
        export_btn.pack(side="left", padx=10)
        
        cancel_btn = RoundedButton(center_frame, text="✕  Cancel", bg="#999999",
                       activebackground="#777777", min_width=100,
                       command=format_window.destroy)
        cancel_btn.pack(side="left", padx=10)

    dl_btn = RoundedButton(action_group, text="⬇  Export Records",
                            bg=C_GREEN,
                            activebackground="#005522",
                            command=export_records)
    dl_btn.pack(side="left", padx=4)

    analytics_btn = RoundedButton(action_group, text="📈  Analytics",
                                   bg="#1A56A0",
                                   activebackground="#0F3870",
                                   min_width=140,
                                   command=lambda: [_cancel_inactivity_timer(),
                                                    show_analytics_screen(root, main_frame, sidebar_nav)])
    analytics_btn.pack(side="left", padx=4)

    # ── TABLE ────────────────────────────────────────────────
    table_outer = tk.Frame(main_frame, bg=C_WHITE,
                            highlightbackground=C_BORDER, highlightthickness=1)
    table_outer.pack(fill="both", expand=True, padx=16, pady=(0, 12))

    columns = ("ID", "Full Name", "Contact", "Address", "Purpose", "Person to See", "Date", "Time")

    style2 = ttk.Style()
    style2.theme_use("clam")
    style2.configure("Modern.Treeview",
                     font=("Segoe UI", 9),
                     rowheight=32,
                     background=C_WHITE,
                     fieldbackground=C_WHITE,
                     foreground=C_TEXT_DARK,
                     borderwidth=0)
    style2.configure("Modern.Treeview.Heading",
                     font=("Segoe UI", 9, "bold"),
                     background=C_RED_DARK,
                     foreground=C_WHITE,
                     relief="flat",
                     borderwidth=0)
    # Make heading text dark gray on hover for better visibility
    try:
        style2.map("Modern.Treeview.Heading",
                   foreground=[('active', C_TEXT_DARK), ('!active', C_WHITE)])
    except Exception:
        pass
    style2.map("Modern.Treeview",
               background=[("selected", C_RED_LIGHT)],
               foreground=[("selected", C_RED_DARK)])

    tree = ttk.Treeview(table_outer, columns=columns,
                        show="headings", selectmode="browse",
                        style="Modern.Treeview")

    widths = {"ID": 45, "Full Name": 160, "Contact": 120,
              "Address": 185, "Purpose": 150, "Person to See": 140, "Date": 95, "Time": 65}
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=widths[col],
                    anchor="center" if col in ("ID", "Date", "Time") else "w",
                    minwidth=widths[col])

    tree.tag_configure("odd",  background="#FFF5F5")
    tree.tag_configure("even", background=C_WHITE)

    scrollbar = ttk.Scrollbar(table_outer, orient="vertical", command=tree.yview)
    h_scrollbar = ttk.Scrollbar(table_outer, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=scrollbar.set, xscrollcommand=h_scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    h_scrollbar.pack(side="bottom", fill="x")
    tree.pack(side="left", fill="both", expand=True)

    def load_table():
        for row in tree.get_children():
            tree.delete(row)
        rows = get_all_visitors(search_var.get().strip())
        for i, row in enumerate(rows):
            tree.insert("", "end", values=row, tags=("even" if i % 2 == 0 else "odd",))
        record_count_var.set(f"{len(rows)} record(s)")

    search_var.trace("w", lambda *_: load_table())
    load_table()

    # ── CHANGE PASSWORD CARD ───────────────────────────────
    pw_section = tk.Frame(main_frame, bg=C_OFF_WHITE)
    pw_section.pack(fill="x", padx=16, pady=(6, 12))

    pw_card = tk.Frame(pw_section, bg=C_WHITE,
                       highlightbackground=C_BORDER, highlightthickness=1)
    pw_card.pack(fill="x")
    tk.Frame(pw_card, bg=C_RED_MID, height=4).pack(fill="x")

    pw_body = tk.Frame(pw_card, bg=C_WHITE)
    pw_body.pack(fill="x", padx=24, pady=16)

    pw_left = tk.Frame(pw_body, bg=C_WHITE)
    pw_left.pack(side="left", fill="y", padx=(0, 30))
    header_row = tk.Frame(pw_left, bg=C_WHITE)
    header_row.pack(fill="x")
    lock_label = load_icon_image_file("lock.png", 16)
    if lock_label:
        icon_lbl = tk.Label(header_row, image=lock_label, bg=C_WHITE)
        icon_lbl._img = lock_label
        icon_lbl.pack(side="left", pady=2)
    tk.Label(header_row, text="Change Admin Password",
             font=("Georgia", 11, "bold"),
             bg=C_WHITE, fg=C_RED_DARK).pack(side="left", padx=(6, 0), pady=2)
    instruction_row = tk.Frame(pw_left, bg=C_WHITE)
    instruction_row.pack(fill="x", pady=(6, 0))
    tk.Label(instruction_row, text="Click the button below to change the admin password.",
             font=("Segoe UI", 8), bg=C_WHITE, fg=C_TEXT_MID).pack(side="left", pady=2)

    pw_right = tk.Frame(pw_body, bg=C_WHITE)
    pw_right.pack(side="left", fill="x", expand=True)

    def open_pw_dialog():
        open_change_password_window(root)

    lock_icon = load_icon_image_file("lock.png", 16)
    upd_btn = RoundedButton(pw_right, text="Change Password",
                            icon=lock_icon, compound="left",
                            bg=C_RED_ACCENT,
                            activebackground=C_RED_DARK,
                            min_width=240,
                            command=open_pw_dialog)
    if lock_icon:
        upd_btn._icon_ref = lock_icon
    upd_btn.pack(anchor="w", pady=(8, 0))


# ============================================================
#  SETTINGS SCREEN  –  Dark Mode
# ============================================================

# Dark mode palette
DK_BG       = "#F5F5F5"
DK_SURFACE  = "#FFFFFF"
DK_SURFACE2 = "#F0F0F0"
DK_BORDER   = "#D0D0D5"
DK_TEXT     = "#1C1C1E"
DK_MUTED    = "#6B6B70"
DK_RED      = "#B22222"
DK_GOLD     = "#C8900A"

def show_settings_screen(root, main_frame, sidebar_nav=None):
    for w in main_frame.winfo_children():
        w.destroy()

    if sidebar_nav:
        sidebar_nav["update"]("settings")

    # Top bar
    topbar = tk.Frame(main_frame, bg=DK_BG, height=56)
    topbar.pack(fill="x")
    topbar.pack_propagate(False)
    tk.Label(topbar, text="  ℹ  ABOUT",
             font=("Georgia", 12, "bold"),
             bg=DK_BG, fg=DK_TEXT).pack(side="left", padx=16, fill="y")
    tk.Frame(main_frame, bg=DK_GOLD, height=3).pack(fill="x")

    # Light scrollable body
    scroll_frame = tk.Frame(main_frame, bg=DK_BG)
    scroll_frame.pack(fill="both", expand=True)

    scrollbar = tk.Scrollbar(scroll_frame, orient="vertical")
    scrollbar.pack(side="right", fill="y")

    body_canvas = tk.Canvas(scroll_frame, bg=DK_BG, highlightthickness=0,
                            yscrollcommand=scrollbar.set)
    body_canvas.pack(side="left", fill="both", expand=True)
    scrollbar.config(command=body_canvas.yview)

    body_inner = tk.Frame(body_canvas, bg=DK_BG)
    inner_id = body_canvas.create_window(0, 0, anchor="nw", window=body_inner)
    body_inner.bind("<Configure>", lambda e: body_canvas.configure(scrollregion=body_canvas.bbox("all")))
    body_canvas.bind("<Configure>", lambda e: body_canvas.itemconfig(inner_id, width=e.width))

    def _on_mousewheel(event):
        body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    body_canvas.bind("<MouseWheel>", _on_mousewheel)
    body_inner.bind("<MouseWheel>", _on_mousewheel)

    def dk_card(parent, title=None, subtitle=None):
        frame = tk.Frame(parent, bg=DK_SURFACE,
                         highlightbackground=DK_BORDER, highlightthickness=1)
        frame.pack(fill="x", padx=40, pady=(24, 0))
        tk.Frame(frame, bg=DK_RED, height=4).pack(fill="x")
        inner = tk.Frame(frame, bg=DK_SURFACE, padx=32, pady=22)
        inner.pack(fill="x")
        if title:
            tk.Label(inner, text=title, font=("Georgia", 13, "bold"),
                     bg=DK_SURFACE, fg=DK_TEXT).pack(anchor="w", pady=(0, 4))
        if subtitle:
            tk.Label(inner, text=subtitle, font=("Segoe UI", 9),
                     bg=DK_SURFACE, fg=DK_MUTED).pack(anchor="w", pady=(0, 14))
            tk.Frame(inner, bg=DK_BORDER, height=1).pack(fill="x", pady=(0, 18))
        frame.bind("<MouseWheel>", _on_mousewheel)
        inner.bind("<MouseWheel>", _on_mousewheel)
        return inner

    def info_row(parent, label, value):
        row = tk.Frame(parent, bg=DK_SURFACE)
        row.pack(fill="x", pady=5)
        lbl1 = tk.Label(row, text=label, font=("Segoe UI", 9, "bold"),
                 bg=DK_SURFACE, fg=DK_MUTED, width=22, anchor="w")
        lbl1.pack(side="left")
        lbl2 = tk.Label(row, text=value, font=("Segoe UI", 9),
                 bg=DK_SURFACE, fg=DK_TEXT)
        lbl2.pack(side="left")
        row.bind("<MouseWheel>", _on_mousewheel)
        lbl1.bind("<MouseWheel>", _on_mousewheel)
        lbl2.bind("<MouseWheel>", _on_mousewheel)

    # ── System Info card ─────────────────────────────────────
    ab = dk_card(body_inner,
                 title="🏛  Barangay San Andres",
                 subtitle="Visitor Log & Records Management System")

    info_row(ab, "📍  Location",    "San Andres, Santiago City, Isabela")
    info_row(ab, "🏢  Office",      "Barangay Records Office")
    info_row(ab, "👤  Managed By",  "Barangay Officials & Staff")
    info_row(ab, "📋  System Name", "Barangay Visitor Log System")
    info_row(ab, "💾  Version",     "2.0")
    info_row(ab, "📅  Year",        "2025")

    # ── Technical card ───────────────────────────────────────
    tech = dk_card(body_inner,
                   title="🛠  Technical Information",
                   subtitle="Stack and environment details")

    info_row(tech, "🐍  Language",   "Python 3")
    info_row(tech, "🖼  UI Library", "Tkinter (built-in)")
    info_row(tech, "🗄  Database",   "SQLite3  —  barangay_visitors.db")
    info_row(tech, "📦  Packaging",  "PyInstaller (for .exe build)")
    info_row(tech, "🖥  Platform",   "Windows / Cross-platform")

    # ── Features card ────────────────────────────────────────
    feat = dk_card(body_inner,
                   title="✅  Features",
                   subtitle="What this system can do")

    features = [
        "Register and log visitor entries",
        "Track purpose of visit and person to see",
        "Admin panel with full visitor records",
        "Search and filter visitor logs",
        "Export records to text file",
        "Analytics dashboard with charts",
        "Auto logout after 5 minutes of inactivity",
        "Password-protected admin access",
    ]
    for feat_text in features:
        row = tk.Frame(feat, bg=DK_SURFACE)
        row.pack(fill="x", pady=3)
        tk.Label(row, text="✦", font=("Segoe UI", 9),
                 bg=DK_SURFACE, fg=DK_GOLD).pack(side="left", padx=(0, 10))
        tk.Label(row, text=feat_text, font=("Segoe UI", 9),
                 bg=DK_SURFACE, fg=DK_TEXT).pack(side="left")

    # ── Developer card ───────────────────────────────────────
    dev = dk_card(body_inner,
                  title="💻  Developer Notes",
                  subtitle="Built for Barangay San Andres")

    note_lines = [
        "This system was developed to digitize and streamline",
        "the barangay's visitor registration process, replacing",
        "manual logbooks with a fast, searchable digital record.",
    ]
    for line in note_lines:
        tk.Label(dev, text=line, font=("Segoe UI", 9),
                 bg=DK_SURFACE, fg=DK_TEXT, anchor="w").pack(fill="x", pady=1)

    # ── Backup Management card ───────────────────────────────
    bk = dk_card(body_inner,
                 title="💾  Database Backup",
                 subtitle="Automatic offline backup to AppData\\Local\\BarangayVisitorLog\\backups")

    # Backup path display
    path_row = tk.Frame(bk, bg=DK_SURFACE)
    path_row.pack(fill="x", pady=(0, 10))
    tk.Label(path_row, text="📂  Backup Folder:",
             font=("Segoe UI", 9, "bold"), bg=DK_SURFACE, fg=DK_MUTED,
             width=18, anchor="w").pack(side="left")
    tk.Label(path_row, text=BACKUP_DIR,
             font=("Segoe UI", 8), bg=DK_SURFACE, fg=DK_TEXT,
             wraplength=480, justify="left").pack(side="left")

    # Status label
    status_var = tk.StringVar(value="")
    status_lbl = tk.Label(bk, textvariable=status_var,
                          font=("Segoe UI", 8, "italic"),
                          bg=DK_SURFACE, fg=C_GREEN, anchor="w")
    status_lbl.pack(fill="x", pady=(0, 10))

    # Buttons row
    btn_row = tk.Frame(bk, bg=DK_SURFACE)
    btn_row.pack(fill="x", pady=(0, 14))

    def do_manual_backup():
        ok, msg = auto_backup_database(silent=True)
        status_var.set("✔ " + msg.split("\n")[0] if ok else "✘ " + msg)
        status_lbl.config(fg=C_GREEN if ok else C_RED_ACCENT)
        refresh_backup_list()

    manual_btn = RoundedButton(btn_row, text="💾  Backup Now",
                               bg=C_RED_ACCENT, activebackground=C_RED_DARK,
                               min_width=160, command=do_manual_backup)
    manual_btn.pack(side="left", padx=(0, 10))

    def open_backup_folder():
        folder = get_backup_dir()
        try:
            os.startfile(folder)
        except Exception:
            import subprocess
            subprocess.Popen(["xdg-open", folder])

    open_btn = RoundedButton(btn_row, text="📂  Open Folder",
                             bg=DK_RED, activebackground=C_RED_DARK,
                             min_width=160, command=open_backup_folder)
    open_btn.pack(side="left")

    # ── Backup list ──────────────────────────────────────────
    tk.Frame(bk, bg=DK_BORDER, height=1).pack(fill="x", pady=(6, 12))
    tk.Label(bk, text="Recent Backups  (max 7 kept)",
             font=("Segoe UI", 9, "bold"), bg=DK_SURFACE, fg=DK_TEXT).pack(anchor="w", pady=(0, 6))

    list_frame = tk.Frame(bk, bg=DK_SURFACE)
    list_frame.pack(fill="x")

    backup_rows = []   # keep refs

    def refresh_backup_list():
        for w in list_frame.winfo_children():
            w.destroy()
        backup_rows.clear()
        backups = get_backup_list()
        if not backups:
            tk.Label(list_frame, text="No backups found yet.",
                     font=("Segoe UI", 8, "italic"),
                     bg=DK_SURFACE, fg=DK_MUTED).pack(anchor="w")
            return
        for i, (fname, fpath, size_kb, modified) in enumerate(backups):
            row_bg = DK_SURFACE if i % 2 == 0 else DK_SURFACE2
            row = tk.Frame(list_frame, bg=row_bg, pady=4)
            row.pack(fill="x")

            info_col = tk.Frame(row, bg=row_bg)
            info_col.pack(side="left", fill="x", expand=True, padx=(4, 0))

            tk.Label(info_col, text=f"  {modified}",
                     font=("Segoe UI", 8, "bold"), bg=row_bg, fg=DK_TEXT).pack(anchor="w")
            tk.Label(info_col, text=f"  {fname}   ({size_kb:.1f} KB)",
                     font=("Segoe UI", 7), bg=row_bg, fg=DK_MUTED).pack(anchor="w")

            restore_btn = RoundedButton(row, text="Restore",
                                        bg=C_RED_DARK, activebackground=C_RED_ACCENT,
                                        font=("Segoe UI", 8, "bold"),
                                        min_width=80,
                                        command=lambda p=fpath: restore_backup(p, parent=root))
            restore_btn.pack(side="right", padx=8)
            backup_rows.append(row)

    refresh_backup_list()

    # Spacer at bottom
    tk.Frame(body_inner, bg=DK_BG, height=40).pack()

    # Bind mousewheel to every widget in the scroll area recursively
    def bind_mousewheel_recursive(widget):
        widget.bind("<MouseWheel>", _on_mousewheel)
        for child in widget.winfo_children():
            bind_mousewheel_recursive(child)
    body_inner.after(100, lambda: bind_mousewheel_recursive(body_inner))


# ============================================================
#  MAIN WINDOW SETUP
# ============================================================

def main():
    create_database()

    root = tk.Tk()
    root.title("Barangay San Andres — Visitor Log System")
    root.geometry("1100x720")
    root.minsize(900, 560)
    root.configure(bg=C_OFF_WHITE)
    configure_button_styles()

    # Set window icon (taskbar + title bar)
    try:
        icon_ico = resource_path(os.path.join("icon", "logo.ico"))
        icon_png = resource_path(os.path.join("logo", "logo.png"))
        if os.path.exists(icon_ico):
            root.iconbitmap(icon_ico)
        elif os.path.exists(icon_png):
            from PIL import Image, ImageTk
            _ico_img = ImageTk.PhotoImage(Image.open(icon_png).resize((32, 32)))
            root.iconphoto(True, _ico_img)
    except Exception:
        pass

    try:
        root.state('zoomed')
    except Exception:
        pass

    # Start periodic auto-backup (every 3 hours, only if DB changed)
    schedule_auto_backup(root)  # runs immediately, then reschedules itself every 3 hours

    # ── SIDEBAR ──────────────────────────────────────────────
    sidebar = tk.Frame(root, bg=C_SIDEBAR_BG, width=210)
    sidebar.pack(side="left", fill="y")
    sidebar.pack_propagate(False)

    # Top accent
    tk.Frame(sidebar, bg=C_GOLD, height=4).pack(fill="x")

    # Logo area
    logo_area = tk.Frame(sidebar, bg=C_SIDEBAR_BG)
    logo_area.pack(fill="x", pady=(20, 0))

    logo_dir = os.path.join(os.path.dirname(__file__), "logo")
    logo_path = os.path.join(logo_dir, "logo.png")
    _sidebar_logos = [None]
    if os.path.exists(logo_path):
        try:
            from PIL import Image, ImageTk
            img = Image.open(logo_path).resize((150, 150), Image.LANCZOS)
            _sidebar_logos[0] = ImageTk.PhotoImage(img)
            lbl = tk.Label(logo_area, image=_sidebar_logos[0], bg=C_SIDEBAR_BG)
            lbl.pack()
            sidebar.logo_img = _sidebar_logos[0]
        except Exception:
            tk.Label(logo_area, text="🏛", font=("Segoe UI", 32),
                     bg=C_SIDEBAR_BG, fg=C_WHITE).pack()
    else:
        tk.Label(logo_area, text="🏛", font=("Segoe UI", 32),
                 bg=C_SIDEBAR_BG, fg=C_WHITE).pack()

    tk.Label(sidebar, text="BARANGAY",
             font=("Georgia", 13, "bold"),
             bg=C_SIDEBAR_BG, fg=C_WHITE).pack(pady=(8, 0))
    tk.Label(sidebar, text="SAN ANDRES",
             font=("Georgia", 13, "bold"),
             bg=C_SIDEBAR_BG, fg=C_WHITE).pack()

    tk.Frame(sidebar, bg=C_GOLD, height=1).pack(fill="x", padx=20, pady=8)

    tk.Label(sidebar, text="V I S I T O R  L O G  S Y S T E M",
             font=("Segoe UI", 7, "bold"),
             bg=C_SIDEBAR_BG, fg=C_TEXT_LIGHT).pack()

    tk.Frame(sidebar, bg=C_RED_MID, height=1).pack(fill="x", padx=0, pady=12)

    # ── NAV ITEMS ─────────────────────────────────────────────
    # We'll store nav frame references for active state management
    nav_frames = {}
    main_frame_widget = None

    def main_frame_ref():
        return main_frame_widget

    sidebar_nav_state = {"current": "visitor_entry"}

    def update_nav(active_key):
        sidebar_nav_state["current"] = active_key
        for key, frame_data in nav_frames.items():
            frame = frame_data["frame"]
            inner = frame_data["inner"]
            is_active = key == active_key
            bg = C_NAV_ACTIVE_BG if is_active else C_SIDEBAR_BG
            frame.config(bg=bg)
            inner.config(bg=bg)
            for w in inner.winfo_children():
                try:
                    w.config(bg=bg)
                    if is_active:
                        w.config(fg=C_WHITE, font=("Segoe UI", 10, "bold"))
                    else:
                        w.config(fg=C_WHITE, font=("Segoe UI", 10))
                except Exception:
                    pass
            # Gold accent bar
            for w in frame.winfo_children():
                if isinstance(w, tk.Frame) and w != inner:
                    w.config(bg=C_GOLD if is_active else C_SIDEBAR_BG)

    nav_icon_images = {
        "visitor_entry": load_icon_image_file("user.png", 16),
        "visitor_log": load_icon_image_file("bullet.png", 16),
        "settings": load_icon_image_file("settings.png", 16),
    }

    nav_items = [
        ("visitor_entry", "👤", "Visitor Entry"),
        ("visitor_log", "•", "Visitor Log"),
        ("settings", "ℹ", "About"),
    ]

    for key, icon, text in nav_items:
        is_active = key == "visitor_entry"
        bg = C_NAV_ACTIVE_BG if is_active else C_SIDEBAR_BG

        frame = tk.Frame(sidebar, bg=bg, cursor="hand2")
        frame.pack(fill="x", pady=1)

        # Gold accent bar (left)
        accent = tk.Frame(frame, bg=C_GOLD if is_active else bg, width=4)
        accent.pack(side="left", fill="y")

        inner = tk.Frame(frame, bg=bg)
        inner.pack(side="left", fill="x", expand=True, padx=10, pady=10)

        icon_img = nav_icon_images.get(key)
        if icon_img:
            icon_lbl = tk.Label(inner, image=icon_img, bg=bg)
            icon_lbl._img = icon_img
        else:
            icon_lbl = tk.Label(inner, text=icon, font=("Segoe UI", 11),
                                bg=bg, fg=C_WHITE)
        icon_lbl.pack(side="left")
        text_lbl = tk.Label(inner, text=f"  {text}",
                            font=("Segoe UI", 10, "bold" if is_active else "normal"),
                            bg=bg, fg=C_WHITE)
        text_lbl.pack(side="left")

        nav_frames[key] = {"frame": frame, "inner": inner, "accent": accent}

        # Hover + click
        def make_handlers(k, f, i, children):
            def on_enter(e):
                if sidebar_nav_state["current"] != k:
                    f.config(bg=C_SIDEBAR_HOVER)
                    i.config(bg=C_SIDEBAR_HOVER)
                    for w in children: w.config(bg=C_SIDEBAR_HOVER)
            def on_leave(e):
                if sidebar_nav_state["current"] != k:
                    f.config(bg=C_SIDEBAR_BG)
                    i.config(bg=C_SIDEBAR_BG)
                    for w in children: w.config(bg=C_SIDEBAR_BG)
            def on_click(e):
                if k == "visitor_entry":
                    show_user_screen(root, mf, {"update": update_nav})
                elif k == "visitor_log":
                    if open_login_window(root):
                        show_admin_screen(root, mf, {"update": update_nav})
                elif k == "settings":
                    show_settings_screen(root, mf, {"update": update_nav})
            return on_enter, on_leave, on_click

        ch = [icon_lbl, text_lbl]
        oe, ol, oc = make_handlers(key, frame, inner, ch)
        for widget in [frame, inner] + ch:
            widget.bind("<Enter>", oe)
            widget.bind("<Leave>", ol)
            widget.bind("<Button-1>", oc)

    # Gold divider
    tk.Frame(sidebar, bg=C_GOLD, height=1).pack(fill="x", padx=0, pady=(20, 8))

    # Live clock
    clock_var = tk.StringVar()
    clock_frame = tk.Frame(sidebar, bg=C_SIDEBAR_BG)
    clock_frame.pack(fill="x", padx=18)

    clock_icon = load_icon_image_file("calendar.png", 16)
    if clock_icon:
        icon_lbl = tk.Label(clock_frame, image=clock_icon, bg=C_SIDEBAR_BG)
        icon_lbl._img = clock_icon
    else:
        icon_lbl = tk.Label(clock_frame, text="📅", font=("Segoe UI", 9),
                            bg=C_SIDEBAR_BG, fg=C_GOLD)
    icon_lbl.pack(side="left", padx=(0, 6))

    clock_lbl = tk.Label(clock_frame, textvariable=clock_var,
                          font=("Segoe UI", 8),
                          bg=C_SIDEBAR_BG, fg=C_TEXT_LIGHT,
                          justify="left")
    clock_lbl.pack(side="left")

    # Store form variables globally so they can be synced with the clock
    form_vars = {"date_var": None, "time_var": None}

    def update_clock():
        now = datetime.now()
        clock_var.set(f"{now.strftime('%A, %B %d, %Y')}\n{now.strftime('%I:%M:%S %p')}")
        
        # Sync form variables if they are set
        if form_vars["date_var"] is not None:
            form_vars["date_var"].set(now.strftime("%Y-%m-%d"))
        if form_vars["time_var"] is not None:
            form_vars["time_var"].set(now.strftime("%I:%M %p"))
            
        root.after(1000, update_clock)
    update_clock()

    # Bottom label
    tk.Frame(sidebar, bg=C_GOLD, height=1).pack(side="bottom", fill="x", pady=(0, 0))
    tk.Label(sidebar, text="Barangay Records Office\nPowered by Python & SQLite",
             font=("Segoe UI", 7), bg=C_SIDEBAR_BG, fg="#A04040",
             justify="center").pack(side="bottom", pady=12)

    # ── MAIN CONTENT ──────────────────────────────────────────
    mf = tk.Frame(root, bg=C_OFF_WHITE)
    mf.pack(side="left", fill="both", expand=True)

    # Share nav updater
    sidebar_nav = {"update": update_nav}
    main_frame_widget = mf
    
    # Attach form_vars to root so it can be accessed from show_user_screen
    root.form_vars = form_vars

    show_user_screen(root, mf, sidebar_nav)
    root.mainloop()

main()